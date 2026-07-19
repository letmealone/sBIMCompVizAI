# -*- coding: utf-8 -*-
"""
ifc_to_excel.py
----------------
임의의 IFC 파일(.ifc)을 입력받아, 건축공학의 일반적 정보 계층
(Project > Site > Building > Storey > Space > Element)에 맞춰
행렬(엔티티 x 속성) 구조의 엑셀 파일로 자동 추출하는 모듈.
"""

import json
import math
import re
from collections import Counter, OrderedDict, defaultdict

import ifcopenshell
import ifcopenshell.util.element as E
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATEGORY_LABELS = {
    'IfcSite': '대지(Site)',
    'IfcBuilding': '건물(Building)',
    'IfcBuildingStorey': '층(Storey)',
    'IfcSpace': '공간(Space)',
    'IfcWall': '벽(Wall)',
    'IfcWallStandardCase': '벽(Wall)',
    'IfcCurtainWall': '커튼월(CurtainWall)',
    'IfcColumn': '기둥(Column)',
    'IfcBeam': '보(Beam)',
    'IfcMember': '부재(Member)',
    'IfcSlab': '바닥(Slab)',
    'IfcCovering': '천장외장(Covering)',
    'IfcRoof': '지붕(Roof)',
    'IfcRailing': '난간(Railing)',
    'IfcStair': '계단(Stair)',
    'IfcStairFlight': '계단참(StairFlight)',
    'IfcRamp': '경사로(Ramp)',
    'IfcRampFlight': '경사로참(RampFlight)',
    'IfcDoor': '문(Door)',
    'IfcWindow': '창(Window)',
    'IfcOpeningElement': '개구부(Opening)',
    'IfcPlate': '플레이트(Plate)',
    'IfcGrid': '그리드(Grid)',
    'IfcFurnishingElement': '적재물(Furnishing)',
    'IfcFlowTerminal': '설비단말(FlowTerminal)',
    'IfcDistributionElement': '설비요소(Distribution)',
    'IfcLightFixture': '조명(LightFixture)',
    'IfcSensor': '센서(Sensor)',
    'IfcFireSuppressionTerminal': '소방장치(FireSuppressionTerminal)',
    'IfcAlarm': '경보기(Alarm)',
}

CORE_ATTRS = ['GlobalId', 'Name', 'Description', 'ObjectType', 'Tag', 'PredefinedType', 'LongName']

EXCLUDE_FROM_WIDE = {'IfcAnnotation'}
MAX_ROWS_LONG = 300_000  

def _safe_attr(ent, name):
    try:
        v = getattr(ent, name)
    except Exception:
        return None
    if v is None:
        return None
    if hasattr(v, 'is_a'):
        try:
            return v.is_a()
        except Exception:
            return str(v)
    return v


def _storey_via_decomposition(ent):
    cur = ent
    seen = set()
    while cur is not None and hasattr(cur, 'is_a'):
        if cur.is_a('IfcBuildingStorey'):
            return cur
        if id(cur) in seen:
            break
        seen.add(id(cur))
        decomposes = getattr(cur, 'Decomposes', None)
        if not decomposes:
            return None
        cur = decomposes[0].RelatingObject
    return None


def _get_storey(ent):
    try:
        c = E.get_container(ent)
    except Exception:
        c = None
    cur = c
    seen = set()
    while cur is not None and hasattr(cur, 'is_a') and not cur.is_a('IfcBuildingStorey'):
        if id(cur) in seen:
            break
        seen.add(id(cur))
        try:
            cur = E.get_container(cur)
        except Exception:
            cur = None
    if cur is None:
        cur = _storey_via_decomposition(ent)
    return cur


def _get_spaces_via_boundary(ifc_file, ent):
    spaces = _get_spaces_list_via_boundary(ifc_file, ent)
    return ', '.join(spaces) if spaces else None


def _get_spaces_list_via_boundary(ifc_file, ent):
    spaces = []
    for rel in ifc_file.by_type('IfcRelSpaceBoundary'):
        if rel.RelatedBuildingElement == ent and rel.RelatingSpace is not None:
            spaces.append(rel.RelatingSpace.Name)
    return spaces


def _get_material(ent):
    try:
        m = E.get_material(ent)
    except Exception:
        return None
    if m is None:
        return None
    try:
        if m.is_a('IfcMaterial'):
            return m.Name
        if m.is_a('IfcMaterialList'):
            return ', '.join([mm.Name for mm in m.Materials])
        if m.is_a() == 'IfcMaterialLayerSetUsage' and m.ForLayerSet:
            return m.ForLayerSet.LayerSetName or 'LayerSet'
        if hasattr(m, 'Name') and m.Name:
            return m.Name
        return m.is_a()
    except Exception:
        return None


def _flatten_psets(ent):
    try:
        psets = E.get_psets(ent, qtos_only=False)
    except Exception:
        psets = {}
    flat = {}
    for pset_name, props in psets.items():
        for k, v in props.items():
            if k == 'id':
                continue
            flat[f"{pset_name}.{k}"] = v
    return flat


def _resolve_body_items(ent):
    if not ent.Representation:
        return []
    items = []
    for rep in ent.Representation.Representations:
        if rep.RepresentationIdentifier != 'Body':
            continue
        for it in rep.Items:
            if it.is_a('IfcMappedItem'):
                mr = it.MappingSource.MappedRepresentation
                items.extend(mr.Items)
            else:
                items.append(it)
    return items


def _profile_xy_extent(profile):
    """[수정사항] IfcIndexedPolyCurve 기반 임의 단면(IfcArbitraryClosedProfileDef 등)은
    벽이 대각선 파사드에 맞춰 비스듬히 잘린 경우 흔한데, 이때 단순 축정렬(axis-aligned)
    bbox를 쓰면 회전된 사각형의 진짜 폭/길이를 제대로 복원하지 못하고 실제보다 훨씬
    작은 값이 나오는 문제가 실측으로 확인됨(예: 실제 길이 1687mm인 벽이 축정렬 bbox로는
    1212mm로 계산되어 약 28% 과소산정). 최소회전사각형(minimum_rotated_rectangle)의 변
    길이를 쓰면 회전 각도와 무관하게 원래 치수를 훨씬 정확히 복원한다(같은 사례에서
    1694mm로 오차 0.4% 이내)."""
    try:
        if profile.is_a('IfcRectangleProfileDef'):
            return profile.XDim, profile.YDim
        if profile.is_a('IfcCircleProfileDef'):
            return profile.Radius * 2, profile.Radius * 2
        curve = getattr(profile, 'OuterCurve', None)
        if curve is not None and curve.is_a('IfcIndexedPolyCurve'):
            pts = curve.Points
            if pts.is_a('IfcCartesianPointList2D'):
                arr = np.array(pts.CoordList, dtype=float)
                try:
                    from shapely.geometry import Polygon
                    poly = Polygon(arr)
                    mrr = poly.minimum_rotated_rectangle
                    coords = list(mrr.exterior.coords)
                    side_a = np.linalg.norm(np.array(coords[1]) - np.array(coords[0]))
                    side_b = np.linalg.norm(np.array(coords[2]) - np.array(coords[1]))
                    if side_a > 0 and side_b > 0:
                        # [수정사항] MRR의 변 순서는 회전각과 무관하게 임의의 시작
                        # 코너부터 매겨지므로, 도형이 이미 축정렬(회전 없음)이어도
                        # side_a가 로컬 X축이 아니라 Y축 방향 길이가 되는 경우가 있었음
                        # (실측 확인: L자 코너 벽 조각에서 X/Y가 뒤바뀌어 이후 Position
                        # 변환 단계에서 코너 위치가 어긋남). 첫 변의 방향이 X축에 더
                        # 가까우면 side_a=X, 아니면 side_a=Y로 판단해 배정한다.
                        edge_vec = np.array(coords[1]) - np.array(coords[0])
                        if abs(edge_vec[0]) >= abs(edge_vec[1]):
                            return float(side_a), float(side_b)
                        else:
                            return float(side_b), float(side_a)
                except Exception:
                    pass
                ext = arr.max(axis=0) - arr.min(axis=0)
                return float(ext[0]), float(ext[1])
    except Exception:
        pass
    return None, None


def _item_extent_corners(it, fallback_rotation=None):
    """Body 아이템 하나의 로컬 좌표계(엔티티 자체 기준) 코너/정점 좌표들을 반환한다.
    여러 아이템(레이어·하드웨어 부품 등)을 하나로 합쳐 전체 bbox를 구할 때 쓰인다.
    [수정사항 1] IfcExtrudedAreaSolid의 프로파일은 자신의 Position(엔티티 로컬 프레임
    내에서의 오프셋+회전)을 기준으로 배치되는데, 예전엔 이를 무시하고 모든 압출체가
    원점(0,0,0)에서 시작한다고 가정했음. 벽 하나가 서로 수직인 압출체 2개(예: L자
    코너를 이루는 두 조각)로 구성된 경우, Position을 무시하면 두 조각이 실제로는
    겹치는 위치에 있는데도 서로 다른 원점 기준으로 합쳐져 폭x폭이 곱해진 것처럼
    부풀려진 bbox가 나오는 문제가 실측으로 확인됨(21m 벽 하나가 450㎡로 과대산정).
    [수정사항 2] 일부 IFC는 압출체 2개 중 하나의 Position에 회전축(Axis)과 기준방향
    (RefDirection)이 둘 다 정의되지 않은(NULL) 경우가 있는데, 이때 회전이 없는(단위
    행렬) 것으로 처리하면 - 같은 형태를 다른 로컬 축 관례로 표현한 형제 아이템(예:
    프로파일이 (10,200)/(200,10)로 서로 축만 바뀐 사실상 동일 형태)과 정렬이 안 맞아
    결합 bbox가 실제보다 훨씬 크게 부풀려지는 문제가 실측 확인됨(실제 200mm 폭인 벽이
    1300mm로 계산됨). 회전 정보가 없는 아이템은, 같은 엔티티의 다른 아이템에서 이미
    확인된 회전(fallback_rotation)을 대신 상속해 이 문제를 피한다 - 같은 벽체를 이루는
    레이어/조각들은 보통 동일한 전체 회전을 공유하고 있을 가능성이 높기 때문이다.
    반환: (코너점 목록, 이 아이템에서 실제로 사용한 회전행렬 또는 None[회전정보 없었음])"""
    if it.is_a('IfcExtrudedAreaSolid'):
        x, y = _profile_xy_extent(it.SweptArea)
        if x is None:
            return [], None
        depth = it.Depth
        # IfcRectangleProfileDef 등 프로파일은 자신의 Position 원점에 대해 대칭 배치되는
        # 것이 IFC 표준 관례
        local_corners = [(sx, sy, sz) for sx in (-x / 2, x / 2) for sy in (-y / 2, y / 2) for sz in (0, depth)]

        pos = it.Position
        has_rotation_info = pos is not None and (pos.Axis is not None or pos.RefDirection is not None)
        rot_used = None
        try:
            import ifcopenshell.util.placement as plc
            if pos is None:
                m = np.eye(4)
            elif has_rotation_info:
                m = plc.get_axis2placement(pos)
                rot_used = m
            elif fallback_rotation is not None:
                # 회전정보 없음 -> 형제 아이템의 회전을 상속하고, 이 아이템 자신의 위치
                # (Location)만 그대로 사용
                m = fallback_rotation.copy()
                m[:3, 3] = np.array(plc.get_axis2placement(pos))[:3, 3]
                rot_used = fallback_rotation
            else:
                m = plc.get_axis2placement(pos) if pos else np.eye(4)
        except Exception:
            m = np.eye(4)
        pts = []
        for (lx, ly, lz) in local_corners:
            p = m @ np.array([lx, ly, lz, 1.0])
            pts.append((p[0], p[1], p[2]))
        return pts, rot_used
    if it.is_a('IfcPolygonalFaceSet') or it.is_a('IfcTriangulatedFaceSet'):
        try:
            return [tuple(p) for p in it.Coordinates.CoordList], None
        except Exception:
            return [], None
    return [], None


def _get_local_dimensions(ent):
    """엔티티의 로컬 X/Y/Z 치수를 구한다.
    [수정사항] 기존에는 Body의 여러 아이템 중 첫 번째로 발견된 것의 bbox만 반환했음
    (예: 문에 손잡이·힌지 등 소형 하드웨어가 별도 IfcPolygonalFaceSet으로 먼저 나열된
    경우, 그 하드웨어 크기가 문 전체 치수로 잘못 채택됨). 이제 Body의 모든 아이템(압출체+
    메쉬)의 코너/정점을 하나로 합쳐 전체를 아우르는 bbox를 계산한다."""
    if ent.Representation:
        for rep in ent.Representation.Representations:
            if rep.RepresentationIdentifier == 'Box':
                for it in rep.Items:
                    if it.is_a('IfcBoundingBox'):
                        return it.XDim, it.YDim, it.ZDim, 'BoundingBox(원본기록값)'

    items = _resolve_body_items(ent)
    all_pts = []
    n_used = 0
    fallback_rotation = None
    for it in items:
        pts, rot_used = _item_extent_corners(it, fallback_rotation=fallback_rotation)
        if rot_used is not None:
            fallback_rotation = rot_used
        if pts:
            all_pts.extend(pts)
            n_used += 1

    if not all_pts:
        return None, None, None, None

    arr = np.array(all_pts, dtype=float)
    ext = arr.max(axis=0) - arr.min(axis=0)
    src = (f'Body 아이템 {n_used}개 통합bbox(직접계산)' if n_used > 1
           else '단일 body 아이템 bbox(직접계산)')
    return float(ext[0]), float(ext[1]), float(ext[2]), src


def _dimension_columns(ent, length_unit_scale=0.001):
    x, y, z, src = _get_local_dimensions(ent)
    conv = lambda v: round(v * length_unit_scale, 4) if v is not None else None
    return {
        '치수_X(m)': conv(x),
        '치수_Y(m)': conv(y),
        '치수_Z(m)': conv(z),
        '치수산출방식': src,
    }

DIMENSION_TARGET_CLASSES = {'IfcWall', 'IfcWallStandardCase', 'IfcColumn', 'IfcBeam', 'IfcMember', 'IfcCurtainWall'}


def _polyline_points_2d(curve):
    try:
        if curve.is_a('IfcIndexedPolyCurve'):
            pts = curve.Points
            if pts.is_a('IfcCartesianPointList2D'):
                return [tuple(p) for p in pts.CoordList]
        elif curve.is_a('IfcPolyline'):
            return [tuple(p.Coordinates) for p in curve.Points]
    except Exception:
        pass
    return None


def _shoelace_area(points):
    n = len(points)
    if n < 3:
        return None
    area = 0.0
    for i in range(n):
        x1, y1 = points[i][0], points[i][1]
        x2, y2 = points[(i + 1) % n][0], points[(i + 1) % n][1]
        area += x1 * y2 - x2 * y1
    return abs(area) / 2.0


def _profile_area(profile):
    try:
        if profile.is_a('IfcRectangleProfileDef'):
            return profile.XDim * profile.YDim
        if profile.is_a('IfcCircleProfileDef'):
            return math.pi * profile.Radius ** 2
        outer = getattr(profile, 'OuterCurve', None)
        if outer is None:
            return None
        pts = _polyline_points_2d(outer)
        if not pts:
            return None
        area = _shoelace_area(pts)
        if area is None:
            return None
        if profile.is_a('IfcArbitraryProfileDefWithVoids'):
            for inner in profile.InnerCurves:
                ipts = _polyline_points_2d(inner)
                if ipts:
                    iarea = _shoelace_area(ipts)
                    if iarea:
                        area -= iarea
        return area
    except Exception:
        return None


def _mesh_surface_area(coords, faces_idx):
    arr = np.array(coords, dtype=float)
    total = 0.0
    for idx in faces_idx:
        if len(idx) < 3:
            continue
        p0 = arr[idx[0]]
        for k in range(1, len(idx) - 1):
            p1, p2 = arr[idx[k]], arr[idx[k + 1]]
            total += float(np.linalg.norm(np.cross(p1 - p0, p2 - p0))) / 2.0
    return total


def _get_footprint_area(ent):
    items = _resolve_body_items(ent)

    for it in items:
        if it.is_a('IfcExtrudedAreaSolid'):
            area = _profile_area(it.SweptArea)
            if area is not None:
                return area, '단면좌표(신발끈공식) 직접계산'

    for it in items:
        if it.is_a('IfcPolygonalFaceSet'):
            coords = it.Coordinates.CoordList
            faces_idx = []
            for face in it.Faces:
                ci = face.CoordIndex
                faces_idx.append(tuple(i - 1 for i in ci))
            if faces_idx:
                return _mesh_surface_area(coords, faces_idx), '메쉬삼각형합산 직접계산'
        if it.is_a('IfcTriangulatedFaceSet'):
            coords = it.Coordinates.CoordList
            faces_idx = [tuple(i - 1 for i in tri) for tri in it.CoordIndex]
            if faces_idx:
                return _mesh_surface_area(coords, faces_idx), '메쉬삼각형합산 직접계산'

    return None, None


_AREA_KEY_RE = re.compile(r'area', re.IGNORECASE)
_RATIO_KEY_RE = re.compile(r'ratio', re.IGNORECASE)


def _get_pset_area_fallback(flat_props):
    """좌표 기반 계산이 모두 실패한 경우에 한해 참고하는 최후의 폴백.
    [수정사항] Qto 속성은 신뢰도 문제로 완전히 배제하고, 그 외 비표준 Pset의
    면적성 속성만 후보로 삼는다."""
    candidates = []
    for key, val in flat_props.items():
        if not isinstance(val, (int, float)):
            continue
        if key.startswith('Qto_'):
            continue
        if not _AREA_KEY_RE.search(key) or _RATIO_KEY_RE.search(key):
            continue
        candidates.append((0 if 'Gross' in key else 1, key, val))
    if not candidates:
        return None, None
    candidates.sort(key=lambda c: c[0])
    _, key, val = candidates[0]
    return val, key


SYSTEM_ASSEMBLY_CLASSES = {'IfcDoor', 'IfcWindow', 'IfcCurtainWall'}

def _get_wall_height_length_mm(ent):
    """벽 하나의 (높이, 길이)를 mm 단위로 분리 반환한다. floorplan_core._get_wall_side_area_m2
    와 동일한 '층고 범위(1.5~8m) 인식' 규칙을 쓴다 - 벽 조립체의 대표 높이를 정할 때
    재사용하기 위해 별도 함수로 분리."""
    x, y, z, src = _get_local_dimensions(ent)
    dims_list = [d for d in (x, y, z) if d is not None]
    if len(dims_list) < 2:
        return None, None
    plausible_height = [d for d in dims_list if 1500.0 <= d <= 8000.0]
    if plausible_height:
        height = max(plausible_height)
        remaining = list(dims_list)
        remaining.remove(height)
        length = max(remaining) if remaining else height
    else:
        dims_sorted = sorted(dims_list, reverse=True)
        length, height = dims_sorted[0], dims_sorted[1]
    return height, length


def _get_system_bounding_face_area(ent):
    x, y, z, src = _get_local_dimensions(ent)
    dims = [d for d in (x, y, z) if d is not None]
    if len(dims) < 2:
        return None, None
    dims.sort(reverse=True)
    area = dims[0] * dims[1]
    return area, f'시스템 전체 bounding치수 기반(폭x높이, 두께제외; {src})'


def _get_decomposed_children(ent, _depth=0, _max_depth=3):
    if _depth >= _max_depth:
        return []
    children = []
    for rel in (getattr(ent, 'IsDecomposedBy', None) or []):
        for child in rel.RelatedObjects:
            children.append(child)
            children.extend(_get_decomposed_children(child, _depth + 1, _max_depth))
    return children


def _get_assembly_bounding_face_area(ent):
    import ifcopenshell.util.placement as plc

    children = _get_decomposed_children(ent)
    if not children:
        return None, None

    all_corners = []
    for child in children:
        x, y, z, _src = _get_local_dimensions(child)
        if x is None or y is None or z is None:
            continue
        if child.ObjectPlacement is None:
            continue
        try:
            m = plc.get_local_placement(child.ObjectPlacement)
        except Exception:
            continue
        for sx in (0, x):
            for sy in (0, y):
                for sz in (0, z):
                    world_pt = m @ np.array([sx, sy, sz, 1.0])
                    all_corners.append(world_pt[:3])

    if len(all_corners) < 2:
        return None, None
    arr = np.array(all_corners)
    extent = arr.max(axis=0) - arr.min(axis=0)
    dims = sorted(extent, reverse=True)
    area = float(dims[0] * dims[1])
    return area, f'하위부품{len(children)}개 월드bounding 근사(폭x높이, 두께제외)'


def _get_union_footprint_area_m2(ent, tol=0.05):
    try:
        import ifcopenshell.geom as geom
        from shapely.geometry import Polygon
        from shapely.ops import unary_union
    except ImportError:
        return None, None
    
    settings = geom.settings()
    settings.set('use-world-coords', True)
    
    def _get_poly(e):
        try:
            shape = geom.create_shape(settings, e)
            verts = np.array(shape.geometry.verts).reshape(-1, 3)
            faces = np.array(shape.geometry.faces).reshape(-1, 3)
            if len(verts) > 0 and len(faces) > 0:
                zmin = verts[:, 2].min()
                polys = []
                for tri in faces:
                    p = verts[tri]
                    if np.all(p[:, 2] <= zmin + tol):
                        try:
                            poly = Polygon(p[:, :2])
                            if poly.is_valid and poly.area > 1e-9:
                                polys.append(poly)
                        except Exception:
                            continue
                if polys:
                    u = unary_union(polys)
                    if not u.is_empty:
                        return u
        except Exception:
            pass
        
        polys = []
        for rel in getattr(e, 'IsDecomposedBy', []):
            for child in rel.RelatedObjects:
                cp = _get_poly(child)
                if cp is not None and not cp.is_empty:
                    polys.append(cp)
        if polys:
            try:
                u = unary_union(polys)
                if not u.is_empty:
                    return u
            except Exception:
                pass
        return None

    poly = _get_poly(ent)
    if poly is not None and not poly.is_empty:
        return poly.area, '하위부품 평면(footprint) 합집합 직접계산'
    return None, None


_FOOTPRINT_AREA_CLASSES = {'IfcSlab', 'IfcRoof', 'IfcCovering', 'IfcSpace'}


def _get_footprint_polygon_area_m2(ent):
    """평면형(수평) 요소의 면적을, floorplan_core에 이미 구현되어 있고 검증된
    평면투영 footprint 계산(z=zmin 삼각형만 골라 union, 여러 조각도 안전하게 통합)을
    재사용해 구한다. Qto 속성은 참조하지 않는다."""
    try:
        import floorplan_core as fc
    except ImportError:
        return None, None
    try:
        poly = fc.get_footprint_polygon_cached(ent)
    except Exception:
        return None, None
    if poly is None or poly.is_empty:
        return None, None
    return poly.area, '평면투영 footprint(다중조각 union, 직접계산)'


def _area_columns(ent, flat_props, length_unit_scale=0.001):
    """[수정사항] Qto 속성(Qto_*.Gross_Area 등)은 저작 툴/생성기에 따라 실제 지오메트리와
    괴리되는 사례가 확인되어(예: 동일 바닥면적의 슬래브인데 Qto값이 2배 이상 차이)
    완전히 배제하고, 좌표(지오메트리) 직접계산을 항상 우선 사용한다. 이렇게 하면 서로
    다른 IFC(전문가 저작 vs AI 생성)를 비교할 때 계산 기준이 동일해진다."""
    cls = ent.is_a()

    # 1순위: 평면형 요소(슬래브/지붕/커버링/공간) - 다중조각도 안전한 검증된 footprint 재사용
    if cls in _FOOTPRINT_AREA_CLASSES:
        area_m2, method = _get_footprint_polygon_area_m2(ent)
        if area_m2 is not None:
            return {'면적(㎡)': round(area_m2, 4), '면적산출방식': method}

    # 2순위: 문/창/커튼월 - 전체 bounding치수(폭x높이) 기반
    # [수정사항] 커튼월은 수직 파사드 부재인데, 예전에는 여기 도달하기 전에 아래
    # '하위부품 footprint 합집합'(평면투영) 방식을 먼저 시도했음. 평면투영은 슬래브처럼
    # 수평인 부재에는 맞지만, 커튼월처럼 수직인 조립체에 쓰면 멀리언·프레임의 바닥
    # 단면(길이x두께 수준의 얇은 조각)만 잡혀 실제 유리 파사드 면적(폭x높이)보다
    # 수십~수백 배 작게 나오는 문제가 실측으로 확인됨(예: 실제로는 수~수십㎡인 커튼월이
    # 0.09~0.75㎡로 계산됨). 문/창과 동일하게 bounding치수(폭x높이) 방식을 먼저 시도하고,
    # 그게 실패할 때만(예: 매우 예외적인 수평 채광창 형태) 평면투영을 최후 폴백으로 쓴다.
    area, method = None, None
    if cls in SYSTEM_ASSEMBLY_CLASSES:
        area, method = _get_system_bounding_face_area(ent)
        if area is None:
            area, method = _get_assembly_bounding_face_area(ent)

    if area is None and cls == 'IfcCurtainWall':
        area_m2, method2 = _get_union_footprint_area_m2(ent)
        if area_m2 is not None:
            return {'면적(㎡)': round(area_m2, 4), '면적산출방식': method2 + '(최후 폴백-수평형태 추정)'}

    # 4순위: 최후 폴백 - 단일 body 기준 직접계산(다중조각 미대응, 위 방법들이 모두 실패한
    # 경우에만 도달)
    if area is None:
        area, method = _get_footprint_area(ent)

    # 5순위: 좌표 계산이 전부 실패한 경우에 한해서만 비Qto Pset 속성값 참고
    if area is None:
        area, source_key = _get_pset_area_fallback(flat_props)
        if area is not None:
            method = f'Pset값 사용({source_key}) - 좌표계산 실패로 인한 참고값'
        else:
            method = '산출불가(좌표계산 실패 + Pset값 없음)'
    else:
        area = area * (length_unit_scale ** 2)
        
    return {
        '면적(㎡)': round(area, 4) if area is not None else None,
        '면적산출방식': method,
    }


AREA_TARGET_CLASSES = {'IfcRoof', 'IfcCovering', 'IfcSlab', 'IfcDoor', 'IfcWindow', 'IfcCurtainWall'}



# ===================================================================
# 3-1b. 벽 내/외벽 판정 및 검증 로직 개선 (선 분할 후 판정 방식)
# ===================================================================

# ===================================================================
# 3-1b. 벽 조립체(레이어 병합) 그룹핑 — 재료 레이어별로 분리 모델링된 벽 보정
# ===================================================================
# 마감(석고보드)+코어(콘크리트/조적)+마감(몰탈) 등으로 벽 하나가 여러 IfcWall
# 엔티티로 나뉘어 모델링된 경우, 개별 레이어는 한쪽 공간에만 RelSpaceBoundary가
# 있거나(마감층) 아예 관계가 없어(코어) 기존 로직에서 '외벽'/'판정불가'로 오판될 수
# 있다. 같은 층에서 서로 평행하고 근접(gap_tol 이내)하며 겹치는 벽들을 하나의
# '조립체'로 묶어, 조립체 전체 기준으로 내/외부를 재판정한다.

def _effective_wall_storey(ifc_file, wall, container_storey, fc, boundary_index=None):
    """[수정사항] 벽의 IFC 컨테이너 소속 층과, 이 벽이 실제 RelSpaceBoundary로 접한
    공간들의 소속 층이 다른 원본 데이터 사례가 실측으로 확인됨(벽 컨테이너는 2층인데
    유일한 관계는 1층 공간을 가리킴 - 450㎡짜리 벽 하나가 통째로 엉뚱한 층에 잡히고
    정작 소속 층에서는 누락되어 인접한 두 층의 층단위 집계가 동시에 틀어짐). 관계된
    공간들의 소속 층 중 컨테이너 층과 일치하는 것이 하나라도 있으면 컨테이너 층을
    그대로 신뢰하고, 하나도 없으면(완전히 어긋난 경우만) 관계된 공간들의 다수 층으로
    보정한다."""
    if boundary_index is None:
        boundary_index = fc._get_boundary_index(ifc_file)
    rels = boundary_index.get(wall.GlobalId, [])
    space_storeys = []
    for r in rels:
        sp = r.RelatingSpace
        if sp is None:
            continue
        st = _get_storey(sp)
        if st is not None:
            space_storeys.append(st)
    if not space_storeys:
        return container_storey

    container_guid = container_storey.GlobalId if container_storey else None
    if any(st.GlobalId == container_guid for st in space_storeys):
        return container_storey

    counts = Counter(st.GlobalId for st in space_storeys)
    majority_guid, _ = counts.most_common(1)[0]
    for st in space_storeys:
        if st.GlobalId == majority_guid:
            return st
    return container_storey


def _wall_footprints_by_storey(ifc_file, fc):
    by_storey = defaultdict(list)
    seen = set()
    boundary_index = fc._get_boundary_index(ifc_file)
    walls = list(ifc_file.by_type('IfcWall'))
    for w in walls:
        if w.GlobalId in seen:
            continue
        seen.add(w.GlobalId)
        st = _get_storey(w)
        if st is None:
            continue
        st = _effective_wall_storey(ifc_file, w, st, fc, boundary_index)
        fp = fc.get_footprint_polygon(w)
        if fp is None or fp.is_empty:
            continue
        by_storey[st.GlobalId].append((w, fp))
    return by_storey


def _wall_run_length(fp):
    """벽 footprint의 '긴 방향(런 길이)' 근사치 - bbox의 더 긴 변."""
    minx, miny, maxx, maxy = fp.bounds
    return max(maxx - minx, maxy - miny)


def _edges_parallel_overlap_length(edges1, edges2, max_gap, angle_tol=0.05):
    """두 벽 외곽선 edge 쌍 중 평행(각도오차 angle_tol 이내)하고 수직간격이 max_gap
    이내인 조합들의 겹침 구간 길이 중 최댓값을 반환(없으면 0.0)."""
    best = 0.0
    for (p1, p2) in edges1:
        d1 = np.array([p2[0] - p1[0], p2[1] - p1[1]])
        len1 = np.linalg.norm(d1)
        if len1 < 1e-6:
            continue
        d1n = d1 / len1
        for (q1, q2) in edges2:
            d2 = np.array([q2[0] - q1[0], q2[1] - q1[1]])
            len2 = np.linalg.norm(d2)
            if len2 < 1e-6:
                continue
            d2n = d2 / len2
            cross = abs(d1n[0] * d2n[1] - d1n[1] * d2n[0])
            if cross > angle_tol:
                continue
            v = np.array([q1[0] - p1[0], q1[1] - p1[1]])
            perp = abs(v[0] * d1n[1] - v[1] * d1n[0])
            if perp > max_gap:
                continue
            t_q1 = (q1[0] - p1[0]) * d1n[0] + (q1[1] - p1[1]) * d1n[1]
            t_q2 = (q2[0] - p1[0]) * d1n[0] + (q2[1] - p1[1]) * d1n[1]
            t_min = max(0.0, min(t_q1, t_q2))
            t_max = min(len1, max(t_q1, t_q2))
            overlap = t_max - t_min
            if overlap > best:
                best = overlap
    return best


def _assembly_length_m(union_footprint):
    """벽 조립체(여러 재료 레이어를 합친) union footprint에서 실제 '길이'(벽이 뻗은
    방향의 장변)를 최소회전사각형 기준으로 구한다(m 단위, union_footprint는 이미 m
    단위 좌표). 조립체는 여러 조각이 이어져 약간 어긋난 MultiPolygon일 수 있어 단순
    축정렬 bbox보다 회전에 안전한 이 방식을 쓴다."""
    try:
        mrr = union_footprint.minimum_rotated_rectangle
        coords = list(mrr.exterior.coords)
        side_a = ((coords[1][0] - coords[0][0]) ** 2 + (coords[1][1] - coords[0][1]) ** 2) ** 0.5
        side_b = ((coords[2][0] - coords[1][0]) ** 2 + (coords[2][1] - coords[1][1]) ** 2) ** 0.5
        return max(side_a, side_b)
    except Exception:
        try:
            minx, miny, maxx, maxy = union_footprint.bounds
            return max(maxx - minx, maxy - miny)
        except Exception:
            return None


def _group_wall_assemblies(walls_fp, fc, max_gap=0.1, min_overlap_abs=0.3, min_overlap_ratio=0.6):
    """같은 층 벽들(walls_fp: [(entity, polygon), ...])을 평행+근접+겹침 기준으로
    Union-Find 그룹핑해 물리적으로 하나의 벽체를 이루는 재료 레이어 묶음을 찾는다.

    [수정사항] 겹침 길이를 절대값(예: 0.3m 이상)만으로 판단하면, 벽 하나가 자기 길이의
    일부 구간에서만 우연히 다른(무관한) 벽과 근접·평행한 경우에도 매칭되어 버리고,
    Union-Find는 전이적(transitive)이라 A-B, B-C가 각각 국소적으로만 참이어도 A/B/C가
    통째로 하나의 조립체로 묶여버리는 문제가 있었다(실제로 실물 건물 데이터에서 서로
    다른 두께·용도의 벽 10개 이상이 하나로 묶이는 사례가 확인됨). 이를 방지하기 위해
    겹침 길이가 절대 최소치(min_overlap_abs)를 넘는 것은 물론, 두 벽 중 더 짧은 쪽
    '런 길이'의 min_overlap_ratio(기본 60%) 이상이어야만 매칭으로 인정한다 - 재료
    레이어(마감+코어+마감)는 보통 거의 동일한 길이로 겹쳐 있으므로 이 조건을 만족하지만,
    우연히 스치듯 근접한 무관한 벽은 비율 조건에서 걸러진다. gap 허용치도 실측 사례
    (0.05~0.08m)에 맞춰 0.15m -> 0.1m로 좁혔다."""
    n = len(walls_fp)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    edges_cache = [fc._polygon_edges(fp) for _, fp in walls_fp]
    bounds = [fp.bounds for _, fp in walls_fp]
    run_lengths = [_wall_run_length(fp) for _, fp in walls_fp]

    for i in range(n):
        bi = bounds[i]
        for j in range(i + 1, n):
            if find(i) == find(j):
                continue
            bj = bounds[j]
            if (bi[2] + max_gap < bj[0] or bj[2] + max_gap < bi[0] or
                    bi[3] + max_gap < bj[1] or bj[3] + max_gap < bi[1]):
                continue  # 바운딩박스 broad-phase: 겹칠 가능성 없으면 스킵
            overlap = _edges_parallel_overlap_length(edges_cache[i], edges_cache[j], max_gap)
            if overlap <= 0:
                continue
            required = max(min_overlap_abs, min_overlap_ratio * min(run_lengths[i], run_lengths[j]))
            if overlap >= required:
                union(i, j)

    groups = defaultdict(list)
    for idx in range(n):
        groups[find(idx)].append(idx)

    from shapely.ops import unary_union
    result = []
    for idxs in groups.values():
        guids = {walls_fp[i][0].GlobalId for i in idxs}
        polys = [walls_fp[i][1] for i in idxs]
        try:
            u = unary_union(polys) if len(polys) > 1 else polys[0]
        except Exception:
            u = polys[0]
        result.append({'guids': guids, 'union': u})
    return result


def _assembly_touching_spaces(union_poly, space_items, adjacency_tol=0.15, min_area=0.01):
    """조립체 union polygon이 실제로(면적 기준) 접하는 공간 GlobalId 집합 반환."""
    touched = set()
    buffered = union_poly.buffer(adjacency_tol)
    for sp_guid, sp_fp in space_items:
        if not buffered.intersects(sp_fp):
            continue
        inter = buffered.intersection(sp_fp)
        if inter.area > min_area:
            touched.add(sp_guid)
    return touched


def _determine_wall_classification(ifc_file):
    """
    [수정사항] 선(先) 분할(Segmentation), 후(後) 판정 로직 적용.
    기존에는 벽 전체(IfcWall) 단위로 판정하여 L자 꺾인 벽이 내/외벽 역할을 동시에 
    수행할 때 모순이 발생했음. 이를 해결하기 위해 공간(Space) 단위로 벽체를 가상으로 
    조각(Segment) 낸 뒤, 각 조각이 다른 공간과 교차하는지를 개별 판정함.
    
    반환 형태: 
    {
      (Wall_GlobalId, Space_GlobalId): ('내벽'|'외벽'|'판정불가', '근거'),
      Wall_GlobalId: ('혼합(내/외벽 복합)'|'내벽'|'외벽', '요약 근거') # 엑셀 및 구버전 호환용
    }
    """
    result = {}
    
    try:
        import floorplan_core as fc
    except ImportError:
        fc = None

    if not fc:
        return {}

    # 1. 공간별 바닥 폴리곤(Footprint) 캐싱
    space_fps = {}
    storey_spaces = defaultdict(list)
    for sp in ifc_file.by_type('IfcSpace'):
        fp = fc.get_footprint_polygon(sp)
        if fp and not fp.is_empty:
            space_fps[sp.GlobalId] = fp
            st = _get_storey(sp)
            if st:
                storey_spaces[st.GlobalId].append((sp.GlobalId, fp))

    # 2. 벽-공간 관계 매핑
    wall_space_map = defaultdict(list)
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        w = r.RelatedBuildingElement
        s = r.RelatingSpace
        if w and w.is_a('IfcWall') and s:
            wall_space_map[w.GlobalId].append(s)

    # 3. 조각(Segment) 단위 교차 검사
    for w_guid, spaces in wall_space_map.items():
        w_ent = ifc_file.by_id(w_guid)
        w_fp = fc.get_footprint_polygon(w_ent)
        
        if not w_fp or w_fp.is_empty:
            for s in spaces:
                result[(w_guid, s.GlobalId)] = ('판정불가', '벽체 지오메트리 추출 실패')
            continue

        for s in spaces:
            s_fp = space_fps.get(s.GlobalId)
            if not s_fp:
                result[(w_guid, s.GlobalId)] = ('판정불가', '공간 지오메트리 추출 실패')
                continue
            
            # [핵심 로직] 공간 영역에 맞게 벽체를 분할(Segmentation)하여 해당 공간과 맞닿은 조각 획득
            seg_result = fc.get_space_wall_segment_polygon(ifc_file, w_ent, s, w_fp)
            if not seg_result or not seg_result[0]:
                result[(w_guid, s.GlobalId)] = ('외벽(추정)', '벽체 분할(Segmentation) 실패로 기본 외벽 간주')
                continue
            
            seg_poly, _method = seg_result
            
            # 분할된 벽체 조각이 반대편에 있는 다른 공간과 닿아있는지 확인 (두께나 모델링 이격 극복을 위해 10cm 버퍼)
            buffered_seg = seg_poly.buffer(0.1)
            
            st = _get_storey(s)
            other_spaces = storey_spaces.get(st.GlobalId, []) if st else []
            
            is_internal = False
            for other_guid, other_fp in other_spaces:
                if other_guid == s.GlobalId:
                    continue
                
                # 다른 공간과 교차하는 영역이 유의미한지 검사
                if buffered_seg.intersects(other_fp):
                    inter = buffered_seg.intersection(other_fp)
                    if inter.area > 0.01:  # 단순 선 접촉이 아닌 실제 면적 공유
                        is_internal = True
                        break
                        
            if is_internal:
                result[(w_guid, s.GlobalId)] = ('내벽', '조각 단위 교차 검사: 반대편에 다른 공간 존재함')
            else:
                result[(w_guid, s.GlobalId)] = ('외벽', '조각 단위 교차 검사: 반대편이 외부와 접함')

    # 4. 엑셀 출력 및 하위 호환성을 위해 단일 GlobalId에 대한 대표 속성(혼합 포함)도 병기
    for w_guid, spaces in wall_space_map.items():
        internal_cnt = sum(1 for s in spaces if result.get((w_guid, s.GlobalId), ('', ''))[0] == '내벽')
        external_cnt = sum(1 for s in spaces if result.get((w_guid, s.GlobalId), ('', ''))[0] in ('외벽', '외벽(추정)'))
        
        if internal_cnt > 0 and external_cnt > 0:
            result[w_guid] = ('혼합(내/외벽 복합)', '분할 조각 중 내벽과 외벽 속성 혼재')
        elif internal_cnt > 0:
            result[w_guid] = ('내벽', '모든 분할 조각이 내벽으로 판정됨')
        elif external_cnt > 0:
            result[w_guid] = ('외벽', '모든 분할 조각이 외벽으로 판정됨')
        else:
            result[w_guid] = ('판정불가', '근거 없음')

    # 5. [조립체 병합 보정 - 국소화 버전] 재료 레이어별로 분리 모델링된 벽(마감+코어+마감
    #    등)은 개별 레이어의 RelSpaceBoundary가 한쪽 공간에만 있거나(마감층) 아예 없어
    #    (코어) 위 3~4단계에서 '외벽'/'판정불가'로 오판될 수 있다. 같은 층에서 서로
    #    평행·근접(0.1m 이내)하며 겹치는 벽들을 하나의 조립체로 묶되,
    #    [수정사항] 처음에는 "이 벽 멤버(GlobalId) 전체 footprint 주변"을 국소 윈도우로
    #    썼는데, 벽 하나가 원래 16m처럼 길게 이어지며 여러 방을 지나가는 경우(예:
    #    복도벽) 그 벽 "전체"의 footprint 자체가 이미 길어서 버퍼를 씌워도 전혀
    #    국소화되지 않는 문제가 실측으로 확인됨. 올바르게 국소화하려면 "벽 전체"가
    #    아니라 "이 벽이 특정 공간 s와 접하는 세그먼트(조각)" 하나만 윈도우로 잡아야
    #    한다 - 이미 3단계에서 쓰는 get_space_wall_segment_polygon()으로 그 조각을 구할
    #    수 있으므로 이를 재사용한다. 조각 주변 1m를 조립체 union과 교차시켜, 같은
    #    국소 위치에 있는 다른 레이어(마감/코어)까지 포함한 국소 단면만 검사한다.
    walls_by_storey = _wall_footprints_by_storey(ifc_file, fc)
    for st_guid, walls_fp in walls_by_storey.items():
        if len(walls_fp) < 2:
            continue
        space_items = storey_spaces.get(st_guid, [])
        if not space_items:
            continue
        assemblies = _group_wall_assemblies(walls_fp, fc)
        guid_to_asm = {}
        for asm in assemblies:
            if len(asm['guids']) > 1:
                for g in asm['guids']:
                    guid_to_asm[g] = asm
        fp_by_guid = {w.GlobalId: fp for w, fp in walls_fp}

        for w_guid, spaces in wall_space_map.items():
            asm = guid_to_asm.get(w_guid)
            if asm is None:
                continue
            w_fp = fp_by_guid.get(w_guid)
            if w_fp is None or w_fp.is_empty:
                continue
            w_ent = ifc_file.by_id(w_guid)

            for s in spaces:
                key = (w_guid, s.GlobalId)
                if result.get(key, ('', ''))[0] == '내벽':
                    continue
                s_fp = space_fps.get(s.GlobalId)
                if not s_fp:
                    continue

                # 3단계와 동일하게, 이 벽이 '이 공간과 접하는' 국소 세그먼트만 추출
                seg_result = fc.get_space_wall_segment_polygon(ifc_file, w_ent, s, w_fp)
                if not seg_result or not seg_result[0]:
                    continue
                seg_poly, _m = seg_result

                # [수정사항] 세그먼트를 등방(모든 방향 동일)으로 buffer(1.0)하면 복도를 따라
                # 이어지는 옆방(같은 쪽에 나란히 있는 다음 공간)까지 창이 넓어져, 반대편이
                # 아니라 '옆방'을 반대편 공간으로 오인하는 문제가 실측으로 확인됨. 벽의
                # 두께(법선) 방향으로만 확장하는 윈도우를 사용한다.
                local_window = fc.get_local_thickness_window(w_ent, seg_poly)
                try:
                    local_union = asm['union'].intersection(local_window)
                except Exception:
                    continue
                if local_union.is_empty:
                    continue
                # [수정사항] adjacency_tol=0으로 하면, 정상적으로 딱 맞닿아있는(gap=0) 공간도
                # 폴리곤 교집합 면적이 0이 되어(경계선 접촉은 면적이 없음) 아예 못 잡는
                # 문제가 있었다(실측 확인: 정상 케이스도 겹침 0으로 나옴). 반대로 너무 큰
                # 버퍼(0.15)는 옆방까지 잡는 문제가 있었다. 실측 검증 결과 0.05가 두
                # 상황을 모두 만족하는 안전한 값이었다(정상 케이스 겹침 0.17 vs 오검출
                # 케이스 겹침 0.003, 임계값 0.01과 충분한 여유 확보).
                touched_local = _assembly_touching_spaces(local_union, space_items,
                                                           adjacency_tol=0.05, min_area=0.01)
                if len(touched_local) < 2:
                    continue

                reason = (f'조립체 병합(국소구간) 판정: 이 벽이 {s.Name or s.GlobalId}과(와) 접하는 '
                          f'국소 구간을 인접 레이어와 묶어 확인한 결과, 서로 다른 공간 '
                          f'{len(touched_local)}개와 접함을 확인')
                result[key] = ('내벽', reason)
                if result.get(w_guid, ('', ''))[0] != '내벽':
                    result[w_guid] = ('내벽', reason)

                # 같은 국소 구간(local_window)에 겹치는 조립체의 다른 멤버(예: RelSpaceBoundary가
                # 아예 없는 코어층)도 물리적으로 같은 위치이므로 함께 내벽으로 승격
                for other_guid in asm['guids']:
                    if other_guid == w_guid:
                        continue
                    other_fp = fp_by_guid.get(other_guid)
                    if other_fp is None or other_fp.is_empty:
                        continue
                    if other_fp.buffer(0.05).intersects(local_window):
                        if result.get(other_guid, ('', ''))[0] != '내벽':
                            result[other_guid] = ('내벽', reason + ' (동일 국소구간의 조립체 멤버)')

    # 6. [재집계] 5단계에서 일부 pair만 국소적으로 '내벽'으로 갱신되었을 수 있으므로,
    #    전체 GlobalId 대표 라벨(혼합 포함)을 pair 결과 기준으로 다시 계산한다.
    for w_guid, spaces in wall_space_map.items():
        internal_cnt = sum(1 for s in spaces if result.get((w_guid, s.GlobalId), ('', ''))[0] == '내벽')
        external_cnt = sum(1 for s in spaces if result.get((w_guid, s.GlobalId), ('', ''))[0] in ('외벽', '외벽(추정)'))

        if internal_cnt > 0 and external_cnt > 0:
            result[w_guid] = ('혼합(내/외벽 복합)', '분할 조각 중 내벽과 외벽 속성 혼재(조립체 병합 국소판정 반영)')
        elif internal_cnt > 0:
            result[w_guid] = ('내벽', '모든 분할 조각이 내벽으로 판정됨(조립체 병합 국소판정 반영)')
        elif external_cnt > 0:
            result[w_guid] = ('외벽', '모든 분할 조각이 외벽으로 판정됨(조립체 병합 국소판정 반영)')
        # else: 관계 자체가 없던 벽(예: 코어)의 whole-wall 라벨은 5단계에서 이미 결정된 값을 유지

    return result


# ===================================================================
# 3-1c. 벽 외 모든 구조부재 대상 범용 내/외부 판정
# ===================================================================

ELEMENT_CLASSIFICATION_TARGET_CLASSES = (
    'IfcSlab', 'IfcRoof', 'IfcCovering',
    'IfcColumn', 'IfcBeam', 'IfcMember', 'IfcCurtainWall', 'IfcDoor', 'IfcWindow',
    'IfcRailing', 'IfcStair', 'IfcStairFlight', 'IfcRamp', 'IfcRampFlight',
)

def _element_both_sides_space_check(ifc_file, target_classes=None):
    import ifcopenshell.util.placement as plc

    normals = {}
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        if elem is None:
            continue
        if target_classes is not None and elem.is_a() not in target_classes:
            continue
        space = r.RelatingSpace
        cg = r.ConnectionGeometry
        if space is None or cg is None:
            continue
        surf = cg.SurfaceOnRelatingElement
        if surf is None or not surf.is_a('IfcCurveBoundedPlane'):
            continue
        plane = surf.BasisSurface
        try:
            space_m = plc.get_local_placement(space.ObjectPlacement)
            plane_m = plc.get_axis2placement(plane.Position)
        except Exception:
            continue
        world_m = space_m @ plane_m
        normal = world_m[:3, 2]
        nrm = np.linalg.norm(normal)
        if nrm == 0:
            continue
        normals.setdefault(elem.GlobalId, []).append(normal / nrm)

    result = {}
    for gid, ns in normals.items():
        ref = ns[0]
        has_same = any(np.dot(n, ref) > 0.5 for n in ns)
        has_opp = any(np.dot(n, ref) < -0.5 for n in ns)
        result[gid] = has_same and has_opp
    return result


def _element_distinct_space_count(ifc_file, target_classes=None):
    counts = defaultdict(set)
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        if elem is None or r.RelatingSpace is None:
            continue
        if target_classes is not None and elem.is_a() not in target_classes:
            continue
        counts[elem.GlobalId].add(r.RelatingSpace.GlobalId)
    return {gid: len(spaces) for gid, spaces in counts.items()}


def _determine_element_classification(ifc_file, target_classes=ELEMENT_CLASSIFICATION_TARGET_CLASSES):
    both_sides = _element_both_sides_space_check(ifc_file, target_classes)
    distinct_count = _element_distinct_space_count(ifc_file, target_classes)

    result = {}
    for cls in target_classes:
        for ent in ifc_file.by_type(cls):
            gid = ent.GlobalId
            if gid in result:
                continue 
            
            if gid in both_sides:
                if both_sides[gid]:
                    result[gid] = ('내부', '2차: RelSpaceBoundary 양면 Space 접촉 확인')
                else:
                    result[gid] = ('외부(추정)', '2차: 한쪽 면만 Space 접촉 → 외부로 추정')
            elif distinct_count.get(gid, 0) >= 2:
                n = distinct_count[gid]
                result[gid] = ('내부(추정-관계기반)', f'3차: 서로 다른 Space {n}개와 연결')
            else:
                result[gid] = ('판정불가', '1차/2차/3차 모두 근거 데이터 없음')
    return result


# ===================================================================
# 3-2. 공간(Space) - 부재(Element) 1:1 매칭 시트
# ===================================================================

def _build_space_element_matrix(ifc_file, wall_classification=None):
    if wall_classification is None:
        wall_classification = _determine_wall_classification(ifc_file)
    rows = []
    for rel in ifc_file.by_type('IfcRelSpaceBoundary'):
        sp = rel.RelatingSpace
        elem = rel.RelatedBuildingElement
        if sp is None or elem is None:
            continue
        storey = _storey_via_decomposition(sp) or _get_storey(sp)
        
        # [수정] 엑셀 내보내기 시 (Wall_GUID, Space_GUID) 조합의 세부 판정 결과 기록
        if elem.is_a('IfcWall'):
            cls_result, cls_reason = wall_classification.get((elem.GlobalId, sp.GlobalId), 
                                     wall_classification.get(elem.GlobalId, (None, None)))
        else:
            cls_result, cls_reason = None, None
            
        rows.append({
            '층(Storey)': storey.Name if storey else None,
            '공간(Space)_Name': sp.Name,
            '공간(Space)_GUID': sp.GlobalId,
            '부재_IFC_Class': elem.is_a(),
            '부재_Name': _safe_attr(elem, 'Name'),
            '부재_GUID': elem.GlobalId,
            'PhysicalOrVirtual': rel.PhysicalOrVirtualBoundary,
            'InternalOrExternal': rel.InternalOrExternalBoundary,
            '벽_내외벽_판정(조각단위)': cls_result,
            '벽_판정근거': cls_reason,
            'RelSpaceBoundary_GUID': rel.GlobalId,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['층(Storey)', '공간(Space)_Name', '부재_IFC_Class', '부재_Name']).reset_index(drop=True)
    return df


def _build_wall_classification_sheet(ifc_file, wall_classification=None):
    if wall_classification is None:
        wall_classification = _determine_wall_classification(ifc_file)
    rows = []
    for w in ifc_file.by_type('IfcWall'):
        storey = _get_storey(w)
        # [수정] 혼합 여부가 저장된 단일 GlobalId 키를 불러와 객체(Wall) 기준의 대푯값 기록
        result, reason = wall_classification.get(w.GlobalId, ('판정불가', '근거 데이터 없음'))
        rows.append({
            '층(Storey)': storey.Name if storey else None,
            'Name': _safe_attr(w, 'Name'),
            'GlobalId': w.GlobalId,
            '내외벽_판정(객체단위)': result,
            '판정근거': reason,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['층(Storey)', '내외벽_판정(객체단위)', 'Name']).reset_index(drop=True)
    return df


# ===================================================================
# 3-3. 공간(Space)에 매칭되지 않은 부재 시트
# ===================================================================

def _build_relspaceboundary_duplication_report(ifc_file):
    pair_count = defaultdict(int)
    elem_class = {}
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        sp = r.RelatingSpace
        if elem is None or sp is None:
            continue
        pair_count[(sp.GlobalId, elem.GlobalId)] += 1
        elem_class[elem.GlobalId] = elem.is_a()

    raw_by_class = defaultdict(int)
    unique_pairs_by_class = defaultdict(int)
    for (sp_guid, elem_guid), cnt in pair_count.items():
        cls = elem_class[elem_guid]
        raw_by_class[cls] += cnt
        unique_pairs_by_class[cls] += 1

    rows = []
    for cls in sorted(set(raw_by_class) | set(unique_pairs_by_class)):
        raw = raw_by_class[cls]
        uniq = unique_pairs_by_class[cls]
        rows.append({
            'IFC_Class': cls,
            '원시_관계건수': raw,
            '고유_공간-부재쌍_수': uniq,
            '초과분(같은공간에중복연결)': raw - uniq,
            '중복여부': '있음' if raw > uniq else '없음',
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values('초과분(같은공간에중복연결)', ascending=False).reset_index(drop=True)
    return df


def _build_unmatched_elements(ifc_file):
    rsb = ifc_file.by_type('IfcRelSpaceBoundary')
    matched_guids = {rel.RelatedBuildingElement.GlobalId
                      for rel in rsb if rel.RelatedBuildingElement is not None}
    matched_classes = {rel.RelatedBuildingElement.is_a()
                        for rel in rsb if rel.RelatedBuildingElement is not None}

    space_storeys = set()
    for sp in ifc_file.by_type('IfcSpace'):
        st = _storey_via_decomposition(sp) or _get_storey(sp)
        if st:
            space_storeys.add(st.Name)

    rows = []
    for ent in ifc_file.by_type('IfcElement'):
        if ent.GlobalId in matched_guids:
            continue
        cls = ent.is_a()
        storey = _get_storey(ent)
        storey_name = storey.Name if storey else None

        if cls not in matched_classes:
            reason = '①구조상 비대상(이 모델에서 해당 클래스는 RelSpaceBoundary 미사용)'
        elif storey_name not in space_storeys:
            reason = '②해당 층에 Space 없음'
        else:
            reason = '③매칭누락 의심(같은 층에 Space 있는데 비어있음)'

        rows.append({
            '층(Storey)': storey_name,
            'IFC_Class': cls,
            'Name': _safe_attr(ent, 'Name'),
            'GlobalId': ent.GlobalId,
            'PredefinedType': _safe_attr(ent, 'PredefinedType'),
            '미매칭_사유': reason,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['미매칭_사유', '층(Storey)', 'IFC_Class', 'Name']).reset_index(drop=True)
    return df


def _label(cls):
    return CATEGORY_LABELS.get(cls, cls)


# ===================================================================
# 1. 계층 구조 추출 (Project > Site > Building > Storey > Space)
# ===================================================================

def _build_hierarchy(ifc_file):
    rows = []

    def add(level, ifc_class, guid, name, parent_path, extra=''):
        rows.append({
            'Level': level, 'IFC_Class': ifc_class, 'GlobalId': guid,
            'Name': name, '상위경로': parent_path, '비고': extra,
        })

    projects = ifc_file.by_type('IfcProject')
    proj = projects[0] if projects else None
    if proj:
        add(0, 'IfcProject', proj.GlobalId, proj.Name, '')
    proj_name = proj.Name if proj else ''

    sites = ifc_file.by_type('IfcSite')
    if not sites:
        sites = [None]

    for site in sites:
        if site is None:
            buildings = ifc_file.by_type('IfcBuilding')
            path0 = proj_name
        else:
            add(1, 'IfcSite', site.GlobalId, site.Name, proj_name)
            path0 = f"{proj_name} > {site.Name}"
            buildings = []
            for rel in (site.IsDecomposedBy or []):
                buildings += [o for o in rel.RelatedObjects if o.is_a('IfcBuilding')]
            if not buildings:
                buildings = ifc_file.by_type('IfcBuilding')

        for building in buildings:
            path1 = f"{path0} > {building.Name}"
            add(2, 'IfcBuilding', building.GlobalId, building.Name, path0)

            storeys = []
            for rel in (building.IsDecomposedBy or []):
                storeys += [o for o in rel.RelatedObjects if o.is_a('IfcBuildingStorey')]
            storeys.sort(key=lambda s: (s.Elevation if s.Elevation is not None else 0))

            for st in storeys:
                path2 = f"{path1} > {st.Name}"
                add(3, 'IfcBuildingStorey', st.GlobalId, st.Name, path1, extra=f"Elevation={st.Elevation}")

                contained = []
                for rel in (st.ContainsElements or []):
                    contained += list(rel.RelatedElements)
                spaces = [c for c in contained if c.is_a('IfcSpace')]
                others = [c for c in contained if not c.is_a('IfcSpace')]

                for sp in spaces:
                    add(4, 'IfcSpace', sp.GlobalId, sp.Name, path2, extra=(sp.LongName or ''))
                if others:
                    add(4, '(요소 합계)', '', f"{len(others)}개 요소 직접 포함", path2,
                        extra=', '.join(sorted(set(o.is_a() for o in others))))

    return pd.DataFrame(rows)


# ===================================================================
# 2. 전체 속성 Long 포맷 (모든 IfcProduct x 모든 Attribute/Pset/Qto)
# ===================================================================

def _build_long_format(ifc_file):
    rows = []
    products = ifc_file.by_type('IfcProduct')

    for ent in products:
        cls = ent.is_a()
        storey = _get_storey(ent)
        storey_name = storey.Name if storey else None
        space_name = None if cls == 'IfcSpace' else _get_spaces_via_boundary(ifc_file, ent)
        material = _get_material(ent)
        guid = _safe_attr(ent, 'GlobalId')
        name = _safe_attr(ent, 'Name')

        for attr in CORE_ATTRS:
            v = _safe_attr(ent, attr)
            if v is None:
                continue
            rows.append({
                'IFC_Class': cls, 'GlobalId': guid, 'Name': name,
                '층(Storey)': storey_name, '공간(Space)': space_name, '재질(Material)': material,
                '구분': 'Attribute', 'Pset': '-', '속성명': attr, '값': v,
            })

        for k, v in _flatten_psets(ent).items():
            pset_name, prop_name = k.split('.', 1)
            kind = 'Qto' if pset_name.startswith('Qto_') else 'Pset'
            rows.append({
                'IFC_Class': cls, 'GlobalId': guid, 'Name': name,
                '층(Storey)': storey_name, '공간(Space)': space_name, '재질(Material)': material,
                '구분': kind, 'Pset': pset_name, '속성명': prop_name, '값': v,
            })

        if len(rows) > MAX_ROWS_LONG:
            break

    return pd.DataFrame(rows)


# ===================================================================
# 3. 클래스별 Wide(행렬) 포맷 - IFC 안에 실제 존재하는 클래스만 동적 생성
# ===================================================================

def _build_wide_sheets(ifc_file):
    products = ifc_file.by_type('IfcProduct')
    by_class = OrderedDict()
    for ent in products:
        cls = ent.is_a()
        if cls in EXCLUDE_FROM_WIDE:
            continue
        by_class.setdefault(cls, []).append(ent)

    priority = ['IfcSite', 'IfcBuilding', 'IfcBuildingStorey', 'IfcSpace']
    ordered_classes = [c for c in priority if c in by_class]
    ordered_classes += sorted(
        [c for c in by_class if c not in priority],
        key=lambda c: -len(by_class[c])
    )

    sheets = OrderedDict()
    for cls in ordered_classes:
        ents = by_class[cls]
        rows = []
        for ent in ents:
            storey = _get_storey(ent)
            flat = _flatten_psets(ent)
            row = {
                'GlobalId': _safe_attr(ent, 'GlobalId'),
                'Name': _safe_attr(ent, 'Name'),
                'ObjectType': _safe_attr(ent, 'ObjectType'),
                'PredefinedType': _safe_attr(ent, 'PredefinedType'),
                'Tag': _safe_attr(ent, 'Tag'),
                '층(Storey)': storey.Name if storey else None,
                '공간(Space)': None if cls == 'IfcSpace' else _get_spaces_via_boundary(ifc_file, ent),
                '재질(Material)': _get_material(ent),
            }
            if cls in DIMENSION_TARGET_CLASSES:
                row.update(_dimension_columns(ent))
            if cls in AREA_TARGET_CLASSES:
                row.update(_area_columns(ent, flat))
            row.update(flat)
            rows.append(row)
        sheets[_label(cls)] = pd.DataFrame(rows)
    return sheets


def load_interest_set(spec_xlsx_path):
    wb = openpyxl.load_workbook(spec_xlsx_path, data_only=True)
    interest = set()

    if '속성목록' in wb.sheetnames:
        ws = wb['속성목록']
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            i_ns   = header.index('Namespace') + 1
            i_prop = header.index('속성명_IFC') + 1
        except ValueError:
            i_ns, i_prop = None, None

        if i_ns and i_prop:
            for r in range(2, ws.max_row + 1):
                ns   = ws.cell(row=r, column=i_ns).value
                prop = ws.cell(row=r, column=i_prop).value
                if ns and prop:
                    interest.add((str(ns).strip(), str(prop).strip()))
            return interest

    def _parse_token(token):
        token = token.strip()
        if not token or any(x in token for x in ['->', '(', ')']):
            return None
        if token.startswith('IfcRel') and '.' not in token:
            return None
        parts = token.split('.')
        if len(parts) < 2:
            return None
        ns, prop = (parts[-2].strip(), parts[-1].strip()) if len(parts) >= 3 \
                   else (parts[0].strip(), parts[1].strip())
        return (ns, prop) if ns and prop else None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, col_idx = None, None
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and 'IFC' in v and '스키마' in v:
                    header_row, col_idx = r, c
                    break
            if header_row:
                break
        if header_row is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            raw = ws.cell(row=r, column=col_idx).value
            if not raw or not isinstance(raw, str):
                continue
            for eq_part in raw.split('='):
                for slash_part in eq_part.split('/'):
                    result = _parse_token(slash_part.strip())
                    if result:
                        interest.add(result)

    return interest


def _format_props_cell(flat_dict, interest_set=None):
    if not flat_dict:
        return ''

    grouped = OrderedDict()
    for k, v in flat_dict.items():
        pset_name, prop_name = k.split('.', 1)
        grouped.setdefault(pset_name, []).append((prop_name, v))

    if not interest_set:
        lines = []
        for pset_name, props in grouped.items():
            lines.append(f"[{pset_name}]")
            for prop_name, v in props:
                v_str = '' if v is None else str(v)
                lines.append(f"  {prop_name} = {v_str}")
        return '\n'.join(lines)

    black = InlineFont(color='000000')
    red = InlineFont(color='FF0000', b=True)
    blocks = []
    buf = ''  

    def flush_black():
        nonlocal buf
        if buf:
            blocks.append(TextBlock(black, buf))
            buf = ''

    for pset_name, props in grouped.items():
        buf += f"[{pset_name}]\n"
        for prop_name, v in props:
            v_str = '' if v is None else str(v)
            line = f"  {prop_name} = {v_str}\n"
            is_interest = (pset_name, prop_name) in interest_set
            if is_interest:
                flush_black()
                blocks.append(TextBlock(red, line))
            else:
                buf += line
    flush_black()

    if not blocks:
        return ''
    last = blocks[-1]
    if last.text.endswith('\n'):
        blocks[-1] = TextBlock(last.font, last.text[:-1])
    return CellRichText(*blocks)


# ===================================================================
# 3-1. 단일 셀(All-in-one cell) 통합 시트
# ===================================================================

def _build_consolidated_sheet(ifc_file, interest_set=None):
    products = ifc_file.by_type('IfcProduct')
    rows = []
    for ent in products:
        cls = ent.is_a()
        if cls in EXCLUDE_FROM_WIDE:
            continue
        storey = _get_storey(ent)
        flat = _flatten_psets(ent)
        base = {
            'IFC_Class': cls,
            '분류': _label(cls),
            '층(Storey)': storey.Name if storey else None,
            'Name': _safe_attr(ent, 'Name'),
            'GlobalId': _safe_attr(ent, 'GlobalId'),
            'ObjectType': _safe_attr(ent, 'ObjectType'),
            'PredefinedType': _safe_attr(ent, 'PredefinedType'),
            'Tag': _safe_attr(ent, 'Tag'),
            '재질(Material)': _get_material(ent),
        }
        if cls in DIMENSION_TARGET_CLASSES:
            base.update(_dimension_columns(ent))
        if cls in AREA_TARGET_CLASSES:
            base.update(_area_columns(ent, flat))
        base['속성개수'] = len(flat)
        base['전체속성(Pset.속성 = 값)'] = _format_props_cell(flat, interest_set)

        space_list = [] if cls == 'IfcSpace' else _get_spaces_list_via_boundary(ifc_file, ent)
        if not space_list:
            row = dict(base)
            row['공간(Space)'] = None
            rows.append(row)
        else:
            for sp_name in space_list:
                row = dict(base)
                row['공간(Space)'] = sp_name
                rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        priority_order = [
            '층(Storey)', '공간(Space)', 'IFC_Class', '분류', 'Name',
            '치수_X(m)', '치수_Y(m)', '치수_Z(m)', '면적(㎡)', '재질(Material)',
            '속성개수', '전체속성(Pset.속성 = 값)', '치수산출방식', '면적산출방식',
        ]
        col_order = [c for c in priority_order if c in df.columns]
        col_order += [c for c in df.columns if c not in col_order]  
        df = df[col_order]
    return df


# ===================================================================
# 4. 엑셀 작성 (서식 포함)
# ===================================================================

HEADER_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
BODY_FONT = Font(name='맑은 고딕', size=10)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_df(ws, df):
    if df.empty:
        ws.append(['(데이터 없음)'])
        return
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for _, r in df.iterrows():
        vals = []
        for v in r.tolist():
            if isinstance(v, (list, tuple)) and not isinstance(v, CellRichText):
                v = ', '.join(str(x) for x in v)
            vals.append(v)
        ws.append(vals)
    for rr in range(2, ws.max_row + 1):
        max_lines = 1
        for cc in range(1, len(df.columns) + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.font = BODY_FONT
            cell.border = BORDER
            n_lines = str(cell.value).count('\n') + 1 if cell.value else 1
            if n_lines > 1:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                max_lines = max(max_lines, n_lines)
            else:
                cell.alignment = Alignment(vertical='center')
        if max_lines > 1:
            ws.row_dimensions[rr].height = min(max_lines * 14, 400)
    sample = df.iloc[:200] if len(df) > 200 else df
    for c in range(1, len(df.columns) + 1):
        col_name = str(df.columns[c - 1])
        if '전체속성' in col_name or '속성정보' in col_name:
            ws.column_dimensions[get_column_letter(c)].width = 80
            continue
        maxlen = max(
            [len(str(df.columns[c - 1]))] +
            [len(str(v).split('\n')[0]) for v in sample.iloc[:, c - 1].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 2, 10), 45)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _unique_sheet_name(wb, name, used):
    base = name[:31]
    candidate = base
    i = 2
    while candidate in used:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


# ===================================================================
# 메인 함수
# ===================================================================

def extract_ifc_to_excel(ifc_path: str, output_path: str = None, include_long: bool = True,
                          consolidated: bool = True, per_class_wide: bool = True,
                          space_element_matrix: bool = True, spec_path: str = None,
                          status_cb=None) -> str:
    if output_path is None:
        output_path = ifc_path.rsplit('.', 1)[0] + '_추출.xlsx'

    ifc_file = ifcopenshell.open(ifc_path)

    interest_set = load_interest_set(spec_path) if spec_path else None

    cnt = Counter(e.is_a() for e in ifc_file)
    df_summary = pd.DataFrame(sorted(cnt.items(), key=lambda x: -x[1]), columns=['IFC_Class', '개수'])
    df_summary.loc[len(df_summary)] = ['IFC Schema', ifc_file.schema]
    if interest_set is not None:
        df_summary.loc[len(df_summary)] = ['관심속성(spec) 매칭개수', len(interest_set)]

    df_hierarchy = _build_hierarchy(ifc_file)

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    ws = wb.create_sheet(_unique_sheet_name(wb, '00_요약', used_names))
    _write_df(ws, df_summary)

    ws = wb.create_sheet(_unique_sheet_name(wb, '01_계층구조', used_names))
    _write_df(ws, df_hierarchy)

    if consolidated:
        df_consol = _build_consolidated_sheet(ifc_file, interest_set)
        ws = wb.create_sheet(_unique_sheet_name(wb, '02_통합조회(단일셀)', used_names))
        _write_df(ws, df_consol)

    if include_long:
        df_long = _build_long_format(ifc_file)
        ws = wb.create_sheet(_unique_sheet_name(wb, '03_전체속성(Long)', used_names))
        _write_df(ws, df_long)

    if space_element_matrix:
        wall_classification = _determine_wall_classification(ifc_file)

        df_matrix = _build_space_element_matrix(ifc_file, wall_classification=wall_classification)
        ws = wb.create_sheet(_unique_sheet_name(wb, '04_공간-부재_매칭(1대1)', used_names))
        _write_df(ws, df_matrix)

        df_unmatched = _build_unmatched_elements(ifc_file)
        ws = wb.create_sheet(_unique_sheet_name(wb, '05_미매칭부재(Space없음)', used_names))
        _write_df(ws, df_unmatched)

        df_wall_class = _build_wall_classification_sheet(ifc_file, wall_classification=wall_classification)
        ws = wb.create_sheet(_unique_sheet_name(wb, '06_벽_내외벽_판정', used_names))
        _write_df(ws, df_wall_class)

        df_dup = _build_relspaceboundary_duplication_report(ifc_file)
        ws = wb.create_sheet(_unique_sheet_name(wb, '07_RelSpaceBoundary_중복진단', used_names))
        _write_df(ws, df_dup)

    if per_class_wide:
        wide_sheets = _build_wide_sheets(ifc_file)
        for sheet_label, df in wide_sheets.items():
            ws = wb.create_sheet(_unique_sheet_name(wb, sheet_label, used_names))
            _write_df(ws, df)

    wb.save(output_path)
    return output_path