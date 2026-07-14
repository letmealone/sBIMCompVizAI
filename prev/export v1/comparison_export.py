# -*- coding: utf-8 -*-
"""
comparison_export.py
----------------------
두 IFC(전문가 vs AI)의 층/공간 매칭 결과와, 매칭된 공간별 객체(부재) 비교, 미매칭
공간의 객체 요약을 엑셀 워크북(여러 시트)으로 만든다.

설계 원칙:
  - ifc_to_excel.py(전체 부재 추출)와 floorplan_core.py(공간/부재 지오메트리, 매칭)의
    기존 함수를 최대한 재사용하고, 여기서는 그 결과를 "비교용으로 조합"만 한다.
  - 공간 매칭은 화면에 보이는 층 하나가 아니라, 주어진 floor_mapping(층 매핑)에 포함된
    "모든 층 쌍"에 대해 일괄 계산한다 - 층 수가 많으면 지오메트리 계산량이 늘어나므로
    status_cb로 진행상황을 알릴 수 있게 했다.
"""
from collections import Counter

import pandas as pd
from openpyxl import Workbook

import floorplan_core as fc
import ifc_to_excel as ite


def _floor_space_polygons(storey_entity, tol=0.05):
    """해당 층 Space들의 footprint 폴리곤만 가볍게 계산한다(구조요소는 계산하지 않음 -
    전체 층 일괄 비교에서 불필요한 구조요소 지오메트리 계산을 생략해 속도를 높인다).
    반환 형식은 floorplan_core.build_storey_plan_data()의 'spaces' 항목과 동일하다."""
    spaces_raw = fc.get_elements_for_storey(storey_entity, classes={'IfcSpace'})
    spaces = []
    for sp in spaces_raw:
        poly = fc.get_footprint_polygon(sp, tol=tol)
        if poly is None:
            continue
        spaces.append({'guid': sp.GlobalId, 'name': sp.Name or '(이름없음)', 'polygon': poly, 'entity': sp})
    return spaces


def _element_class_counts(ifc_file, space_entity):
    """공간에 접한 부재를 IFC 클래스별 개수로 집계.
    get_space_related_elements()가 이미 GlobalId 기준 중복 제거를 하므로 그대로 재사용."""
    elements = fc.get_space_related_elements(ifc_file, space_entity)
    return Counter(e.is_a() for e in elements)


def build_floor_mapping_df(floor_mapping, is_llm=False, llm_detail=None):
    """01_층_매칭 시트용 DataFrame."""
    detail_by_a = {d['a_name']: d for d in (llm_detail or [])}
    rows = []
    for a_name, b_name in floor_mapping.items():
        row = {'전문가(A) 층': a_name, 'AI(B) 층': b_name or '(대응없음)'}
        if is_llm:
            d = detail_by_a.get(a_name, {})
            row['확신도'] = d.get('confidence', '')
            row['판단근거'] = d.get('reason', '')
        rows.append(row)
    return pd.DataFrame(rows)


def compute_all_space_matches(data_a, data_b, floor_mapping, area_thresh=2.0, centroid_thresh=1.0,
                               status_cb=None):
    """floor_mapping에 포함된 모든 층 쌍에 대해 공간 매칭을 일괄 계산.
    반환: (matched_pairs, unmatched_a, unmatched_b, df_space_matching)
      - matched_pairs: [{'A_층','A_공간','A_면적(㎡)','B_층','B_공간','B_면적(㎡)',
                          '면적차(㎡)','centroid거리(m)','A_entity','B_entity'}, ...]
      - unmatched_a/b: [(층이름, space_entry), ...] (space_entry는 _floor_space_polygons 형식)
    """
    matched_pairs, unmatched_a, unmatched_b = [], [], []
    storeys_b_by_name = {s['Name']: s for s in data_b['storeys']}

    for a_storey in data_a['storeys']:
        a_name = a_storey['Name']
        b_name = floor_mapping.get(a_name)
        if status_cb:
            status_cb(f'공간 매칭 계산 중: {a_name} ...')
        spaces_a = _floor_space_polygons(a_storey)

        if not b_name or b_name not in storeys_b_by_name:
            for sp in spaces_a:
                unmatched_a.append((a_name, sp))
            continue

        b_storey = storeys_b_by_name[b_name]
        spaces_b = _floor_space_polygons(b_storey)
        a_to_b, b_to_a, offset, match_info = fc.match_spaces(
            spaces_a, spaces_b, area_thresh=area_thresh, centroid_thresh=centroid_thresh)

        by_guid_a = {s['guid']: s for s in spaces_a}
        by_guid_b = {s['guid']: s for s in spaces_b}
        for m in match_info:
            sa, sb = by_guid_a[m['a_guid']], by_guid_b[m['b_guid']]
            matched_pairs.append({
                'A_층': a_name, 'A_공간': sa['name'], 'A_면적(㎡)': round(sa['polygon'].area, 2),
                'B_층': b_name, 'B_공간': sb['name'], 'B_면적(㎡)': round(sb['polygon'].area, 2),
                '면적차(㎡)': m['area_diff_m2'], 'centroid거리(m)': m['centroid_dist_m'],
                'A_entity': sa['entity'], 'B_entity': sb['entity'],
            })
        for sp in spaces_a:
            if sp['guid'] not in a_to_b:
                unmatched_a.append((a_name, sp))
        for sp in spaces_b:
            if sp['guid'] not in b_to_a:
                unmatched_b.append((b_name, sp))

    df_space = pd.DataFrame([
        {k: v for k, v in r.items() if k not in ('A_entity', 'B_entity')} for r in matched_pairs
    ])
    return matched_pairs, unmatched_a, unmatched_b, df_space


def build_matched_object_comparison_df(ifc_a, ifc_b, matched_pairs):
    """03_매칭공간_객체비교 시트용 DataFrame: 매칭된 공간 쌍마다 접한 부재를
    IFC 클래스별 개수로 비교(A개수/B개수/차이)."""
    rows = []
    for m in matched_pairs:
        cnt_a = _element_class_counts(ifc_a, m['A_entity'])
        cnt_b = _element_class_counts(ifc_b, m['B_entity'])
        row = {
            'A_층': m['A_층'], 'A_공간': m['A_공간'], 'B_층': m['B_층'], 'B_공간': m['B_공간'],
            'A_면적(㎡)': m['A_면적(㎡)'], 'B_면적(㎡)': m['B_면적(㎡)'], '면적차(㎡)': m['면적차(㎡)'],
        }
        for c in sorted(set(cnt_a) | set(cnt_b)):
            a_n, b_n = cnt_a.get(c, 0), cnt_b.get(c, 0)
            row[f'{c}_A개수'] = a_n
            row[f'{c}_B개수'] = b_n
            row[f'{c}_차이(A-B)'] = a_n - b_n
        rows.append(row)
    return pd.DataFrame(rows)


def build_unmatched_object_summary_df(ifc_a, unmatched_a, ifc_b, unmatched_b):
    """04_미매칭공간_객체요약 시트용 DataFrame: 대응 공간을 못 찾은 공간들의 접한
    부재를 클래스별 개수로 요약(어느 쪽 파일 소속인지 '출처' 컬럼으로 구분)."""
    rows = []
    for ifc_file, unmatched_list, side_label in (
        (ifc_a, unmatched_a, '전문가(A)'), (ifc_b, unmatched_b, 'AI(B)'),
    ):
        for storey_name, sp in unmatched_list:
            cnt = _element_class_counts(ifc_file, sp['entity'])
            row = {'출처': side_label, '층': storey_name, '공간': sp['name'],
                   '면적(㎡)': round(sp['polygon'].area, 2)}
            for c, n in cnt.items():
                row[f'{c}_개수'] = n
            rows.append(row)
    return pd.DataFrame(rows)


def build_comparison_workbook(data_a, data_b, floor_mapping, is_llm_floor_mapping=False,
                               llm_floor_detail=None, area_thresh=2.0, centroid_thresh=1.0,
                               status_cb=None):
    """네 시트(층 매칭 / 공간 매칭 / 매칭공간 객체비교 / 미매칭공간 객체요약)로 구성된
    비교 결과 워크북을 생성해 반환한다(저장은 호출부에서)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    if status_cb:
        status_cb('층 매칭 시트 작성 중...')
    df_floor = build_floor_mapping_df(floor_mapping, is_llm=is_llm_floor_mapping, llm_detail=llm_floor_detail)
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '01_층_매칭', used_names))
    ite._write_df(ws, df_floor)

    matched_pairs, unmatched_a, unmatched_b, df_space = compute_all_space_matches(
        data_a, data_b, floor_mapping, area_thresh=area_thresh, centroid_thresh=centroid_thresh,
        status_cb=status_cb,
    )
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '02_공간_매칭', used_names))
    ite._write_df(ws, df_space)

    if status_cb:
        status_cb('매칭된 공간별 객체 비교 계산 중...')
    df_compare = build_matched_object_comparison_df(data_a['ifc_file'], data_b['ifc_file'], matched_pairs)
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '03_매칭공간_객체비교', used_names))
    ite._write_df(ws, df_compare)

    if status_cb:
        status_cb('미매칭 공간 객체 요약 계산 중...')
    df_unmatched = build_unmatched_object_summary_df(data_a['ifc_file'], unmatched_a, data_b['ifc_file'], unmatched_b)
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '04_미매칭공간_객체요약', used_names))
    ite._write_df(ws, df_unmatched)

    return wb
