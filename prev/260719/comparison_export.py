# -*- coding: utf-8 -*-
"""
comparison_export.py
----------------------
두 IFC(전문가 vs AI)의 층/공간 매칭 결과와, 매칭된 공간별 객체(부재) 비교, 미매칭
공간의 객체 요약을 엑셀 워크북(여러 시트)으로 만든다.

시트 구성:
  01_층_매칭            : 층 매핑 테이블
  02_공간_매칭          : 전체 층에 걸친 공간 매핑 테이블 (면적/centroid 오차 포함)
  03_구조부재_내외구분비교 : 벽 포함 모든 구조부재를 내부/외부/판정불가로 나눠 A/B 개수·면적 비교
  03b_구조부재_통합비교  : 같은 구조부재를 내/외부 구분 없이 클래스별로만 A/B 비교 (공간 단위)
  03c_층단위_부재비교    : 03b과 같은 클래스별 개수/면적 비교를 "공간 매칭과 무관하게 층 전체
                          단위"로 집계 (공간 안분/RelSpaceBoundary 계산이 필요 없어 추가 계산
                          비용이 거의 없음 - 층에 속한 부재를 그대로 클래스별 합산만 함)
  04_설비_비교          : 설비(조명/센서/소방장치/경보기 등, RelContainedInSpatialStructure 기반)
                          별도 비교 - 구조부재와 데이터 소스 자체가 달라 항상 분리되어 있었음
  05_미매칭공간_객체요약 : 대응 공간을 못 찾은 공간들의 접한 부재 요약

설계 원칙:
  - ifc_to_excel.py(전체 부재 추출)와 floorplan_core.py(공간/부재 지오메트리, 매칭,
    범용 내/외부 판정·집계)의 기존 함수를 최대한 재사용하고, 여기서는 그 결과를
    "비교용으로 조합"만 한다.
  - 공간 매칭은 화면에 보이는 층 하나가 아니라, 주어진 floor_mapping(층 매핑)에 포함된
    "모든 층 쌍"에 대해 일괄 계산한다. 호출부(Streamlit 앱)가 화면에서 이미 계산해
    캐싱해둔 층별 Space 폴리곤(spaces_dict_a/b)을 넘기면 지오메트리를 다시 계산하지
    않고 재사용한다 - 내보내기가 느려지는 주된 원인이 이 중복 계산이었기 때문.
"""
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import floorplan_core as fc
import ifc_to_excel as ite

# 차이가 나는 셀에 적용할 배경색 (연한 붉은색)
DIFF_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")


def _floor_space_polygons(storey_entity, tol=0.05):
    """해당 층 Space들의 footprint 폴리곤만 가볍게 계산한다(구조요소는 계산하지 않음 -
    전체 층 일괄 비교에서 불필요한 구조요소 지오메트리 계산을 생략해 속도를 높인다).
    반환 형식은 floorplan_core.build_storey_plan_data()의 'spaces' 항목과 동일하다.
    (호출부가 화면에서 이미 캐싱한 결과를 갖고 있지 않을 때의 폴백으로만 쓰인다.)"""
    spaces_raw = fc.get_elements_for_storey(storey_entity, classes={'IfcSpace'})
    spaces = []
    for sp in spaces_raw:
        poly = fc.get_footprint_polygon(sp, tol=tol)
        if poly is None:
            continue
        spaces.append({'guid': sp.GlobalId, 'name': sp.Name or '(이름없음)', 'polygon': poly, 'entity': sp})
    return spaces


def build_floor_mapping_df(floor_mapping, is_llm=False, llm_detail=None):
    """01_층_매칭 시트용 DataFrame."""
    detail_by_a = {d['a_name']: d for d in (llm_detail or [])}
    rows = []
    for a_name, b_name in floor_mapping.items():
        row = {'전문가(A) 층': a_name, 'AI(B) 층': b_name or '(대응없음)'}
        if is_llm:
            d = detail_by_a.get(a_name, {})
            row['확신도'] = d.get('confidence', '')
            row['판단근거'] = d.get('reason', '')
        rows.append(row)
    return pd.DataFrame(rows)


def compute_all_space_matches(data_a, data_b, floor_mapping, spaces_dict_a, spaces_dict_b,
                               area_thresh=2.0, centroid_thresh=1.0, status_cb=None):
    """floor_mapping에 포함된 모든 층 쌍에 대해 공간 매칭을 일괄 계산.
    spaces_dict_a/b: {층이름: spaces} - 호출부(Streamlit 앱)가 화면 캐시를 재사용해 미리
    만들어 넘긴다(없는 층은 호출부가 _floor_space_polygons로 채워서 넘겨야 함).

    반환: (matched_pairs, unmatched_a, unmatched_b, df_space_matching)
      - matched_pairs: [{'A_층','A_공간','A_면적(㎡)','B_층','B_공간','B_면적(㎡)',
                          '면적차(㎡)','centroid거리(m)','A_entity','B_entity'}, ...]
      - unmatched_a/b: [(층이름, space_entry), ...]
    """
    matched_pairs, unmatched_a, unmatched_b = [], [], []
    storeys_b_by_name = {s['Name']: s for s in data_b['storeys']}
    total = len(data_a['storeys'])

    for i, a_storey in enumerate(data_a['storeys'], start=1):
        a_name = a_storey['Name']
        b_name = floor_mapping.get(a_name)
        if status_cb:
            status_cb(f'공간 매칭 계산 중: {a_name} ({i}/{total})')

        spaces_a = spaces_dict_a.get(a_name)
        if spaces_a is None:
            spaces_a = _floor_space_polygons(a_storey)

        if not b_name or b_name not in storeys_b_by_name:
            for sp in spaces_a:
                unmatched_a.append((a_name, sp))
            continue

        b_storey = storeys_b_by_name[b_name]
        spaces_b = spaces_dict_b.get(b_name)
        if spaces_b is None:
            spaces_b = _floor_space_polygons(b_storey)

        a_to_b, b_to_a, offset, match_info = fc.match_spaces(
            spaces_a, spaces_b, area_thresh=area_thresh, centroid_thresh=centroid_thresh)

        by_guid_a = {s['guid']: s for s in spaces_a}
        by_guid_b = {s['guid']: s for s in spaces_b}
        for m in match_info:
            sa, sb = by_guid_a[m['a_guid']], by_guid_b[m['b_guid']]
            matched_pairs.append({
                'A_층': a_name, 'A_공간': sa['name'], 'A_면적(㎡)': round(sa['polygon'].area, 2),
                'B_층': b_name, 'B_공간': sb['name'], 'B_면적(㎡)': round(sb['polygon'].area, 2),
                '면적차(㎡)': m['area_diff_m2'], 'centroid거리(m)': m['centroid_dist_m'],
                'A_entity': sa['entity'], 'B_entity': sb['entity'],
            })
        for sp in spaces_a:
            if sp['guid'] not in a_to_b:
                unmatched_a.append((a_name, sp))
        for sp in spaces_b:
            if sp['guid'] not in b_to_a:
                unmatched_b.append((b_name, sp))

    df_space = pd.DataFrame([
        {k: v for k, v in r.items() if k not in ('A_entity', 'B_entity')} for r in matched_pairs
    ])
    return matched_pairs, unmatched_a, unmatched_b, df_space


def _compute_breakdowns(data_a, data_b, matched_pairs, status_cb=None):
    """매칭된 공간 쌍마다 build_space_structural_breakdown을 '한 번만' 계산해 캐싱한다.
    03(내외구분)과 03b(통합) 시트가 이 결과를 공유해서 쓰므로, 같은 계산을 두 번
    반복하지 않는다(이전 버전은 시트마다 따로 계산해 속도가 거의 2배 느렸음)."""
    total = len(matched_pairs)
    out = []
    for i, m in enumerate(matched_pairs, start=1):
        if status_cb and total:
            status_cb(f'구조부재 집계 계산 중: {m["A_공간"]} ({i}/{total})')
        bd_a = fc.build_space_structural_breakdown(
            data_a['ifc_file'], data_a['element_classification'], m['A_entity'])
        bd_b = fc.build_space_structural_breakdown(
            data_b['ifc_file'], data_b['element_classification'], m['B_entity'])
        out.append((m, bd_a, bd_b))
    return out


def _floor_class_summary(ifc_file, storey_dict):
    """해당 층에 '배치된' 전체 구조부재를 클래스별로 집계한다 (공간 매칭/안분과 무관하게
    층 하나에 속한 부재 전체를 그대로 합산 - 계산 비용이 거의 없다: footprint 폴리곤도,
    RelSpaceBoundary 정밀계산도 필요 없음).
    면적 산정은 공간별 집계(build_space_structural_breakdown)와 동일한 우선순위를 따르되
    안분(apportion)은 하지 않는다(층 전체 기준이라 여러 공간에 걸치는지 여부가 무관함):
      - IfcWall/IfcWallStandardCase: Qto_WallBaseQuantities.Gross_Side_Area 전체값.
      - _STRUCTURAL_AREA_MEANINGFUL_CLASSES(Slab/Roof/Covering/Door/Window/CurtainWall):
        ite._area_columns() 전체값(이제 Qto 정식 수량값을 최우선으로 사용함).
      - 그 외 클래스: 개수만.
    반환: {클래스: {'count':int, 'area':float|None}}"""
    elements = fc.get_elements_for_storey(storey_dict, classes=set(ite.ELEMENT_CLASSIFICATION_TARGET_CLASSES))
    summary = defaultdict(lambda: {'count': 0, 'area': 0.0, '_has_area': False})
    for e in elements:
        cls = e.is_a()
        area_val = None
        if cls in ('IfcWall', 'IfcWallStandardCase'):
            flat = ite._flatten_psets(e)
            area_val, _src = fc._get_wall_side_area_m2(e, flat_props=flat)
        elif cls in fc._STRUCTURAL_AREA_MEANINGFUL_CLASSES:
            flat = ite._flatten_psets(e)
            cols = ite._area_columns(e, flat)
            area_val = cols['면적(㎡)']
        summary[cls]['count'] += 1
        if area_val is not None:
            summary[cls]['area'] += area_val
            summary[cls]['_has_area'] = True
    return {c: {'count': v['count'], 'area': round(v['area'], 2) if v['_has_area'] else None}
            for c, v in summary.items()}


def build_floor_level_comparison_df(data_a, data_b, floor_mapping):
    """03c_층단위_부재비교 시트용 DataFrame: 공간 매칭과 무관하게, 매핑된 층 쌍마다
    전체 부재를 클래스별로 집계해 A/B 개수·면적을 비교한다. 이미 로드된 IFC 정보만
    사용하므로(공간 footprint 계산도, RelSpaceBoundary 정밀계산도 필요 없음) 추가
    지오메트리 계산 비용이 거의 없다."""
    rows = []
    storeys_b_by_name = {s['Name']: s for s in data_b['storeys']}
    for a_storey in data_a['storeys']:
        a_name = a_storey['Name']
        b_name = floor_mapping.get(a_name)
        summary_a = _floor_class_summary(data_a['ifc_file'], a_storey)
        summary_b = (_floor_class_summary(data_b['ifc_file'], storeys_b_by_name[b_name])
                     if b_name and b_name in storeys_b_by_name else {})

        for cls in sorted(set(summary_a) | set(summary_b)):
            va = summary_a.get(cls, {'count': 0, 'area': None})
            vb = summary_b.get(cls, {'count': 0, 'area': None})
            rows.append({
                'A_층': a_name, 'B_층': b_name or '(대응없음)', 'IFC_Class': cls,
                'A_개수': va['count'], 'B_개수': vb['count'],
                'A_면적(㎡)': va['area'], 'B_면적(㎡)': vb['area'],
            })
    return pd.DataFrame(rows)


def build_structural_split_df(breakdowns):
    """03_구조부재_내외구분비교 시트용 DataFrame: 매칭된 공간 쌍마다, 벽 포함 모든
    구조부재를 (IFC_Class, 내/외부 구분)별로 나눠 A/B 개수·면적을 비교한다.
    breakdowns: _compute_breakdowns()의 반환값."""
    rows = []
    for m, bd_a, bd_b in breakdowns:
        base = {'A_층': m['A_층'], 'A_공간': m['A_공간'], 'B_층': m['B_층'], 'B_공간': m['B_공간']}

        keys = set()
        for cls, labels in bd_a['by_class_split'].items():
            keys |= {(cls, label) for label in labels}
        for cls, labels in bd_b['by_class_split'].items():
            keys |= {(cls, label) for label in labels}

        for cls, label in sorted(keys):
            va = bd_a['by_class_split'].get(cls, {}).get(label, {'count': 0, 'area': None})
            vb = bd_b['by_class_split'].get(cls, {}).get(label, {'count': 0, 'area': None})
            row = dict(base)
            row.update({
                'IFC_Class': cls, '구분(내/외부)': label,
                'A_개수': va['count'], 'B_개수': vb['count'],
                'A_면적(㎡)': va['area'], 'B_면적(㎡)': vb['area'],
            })
            rows.append(row)
    return pd.DataFrame(rows)


def build_structural_total_df(breakdowns):
    """03b_구조부재_통합비교 시트용 DataFrame: 내/외부 구분 없이 클래스별로만 A/B 비교.
    breakdowns: _compute_breakdowns()의 반환값(03 시트와 동일한 계산결과 재사용)."""
    rows = []
    for m, bd_a, bd_b in breakdowns:
        base = {'A_층': m['A_층'], 'A_공간': m['A_공간'], 'B_층': m['B_층'], 'B_공간': m['B_공간']}

        classes = set(bd_a['by_class_total']) | set(bd_b['by_class_total'])
        for cls in sorted(classes):
            va = bd_a['by_class_total'].get(cls, {'count': 0, 'area': None})
            vb = bd_b['by_class_total'].get(cls, {'count': 0, 'area': None})
            row = dict(base)
            row.update({
                'IFC_Class': cls,
                'A_개수': va['count'], 'B_개수': vb['count'],
                'A_면적(㎡)': va['area'], 'B_면적(㎡)': vb['area'],
            })
            rows.append(row)
    return pd.DataFrame(rows)


def build_equipment_comparison_df(ifc_a, ifc_b, matched_pairs):
    """04_설비_비교 시트용 DataFrame: 설비(EQUIPMENT_CLASSES)는 구조부재와 데이터 소스
    (RelContainedInSpatialStructure)가 달라 항상 별도 시트로 구성한다. 설비는 면적 개념이
    없으므로 개수만 비교."""
    rows = []
    for m in matched_pairs:
        eq_a = Counter(e.is_a() for e in fc.get_space_contained_equipment(ifc_a, m['A_entity']))
        eq_b = Counter(e.is_a() for e in fc.get_space_contained_equipment(ifc_b, m['B_entity']))
        base = {'A_층': m['A_층'], 'A_공간': m['A_공간'], 'B_층': m['B_층'], 'B_공간': m['B_공간']}
        for cls in sorted(set(eq_a) | set(eq_b)):
            row = dict(base)
            row.update({'설비_클래스': cls, 'A_개수': eq_a.get(cls, 0), 'B_개수': eq_b.get(cls, 0)})
            rows.append(row)
    return pd.DataFrame(rows)


def build_unmatched_object_summary_df(ifc_a, unmatched_a, ifc_b, unmatched_b):
    """05_미매칭공간_객체요약 시트용 DataFrame: 대응 공간을 못 찾은 공간들의 접한
    구조부재를 클래스별 개수로 요약(어느 쪽 파일 소속인지 '출처' 컬럼으로 구분)."""
    rows = []
    for ifc_file, unmatched_list, side_label in (
        (ifc_a, unmatched_a, '전문가(A)'), (ifc_b, unmatched_b, 'AI(B)'),
    ):
        for storey_name, sp in unmatched_list:
            cnt = Counter(e.is_a() for e in fc.get_space_related_elements(ifc_file, sp['entity']))
            row = {'출처': side_label, '층': storey_name, '공간': sp['name'],
                   '면적(㎡)': round(sp['polygon'].area, 2)}
            for c, n in cnt.items():
                row[f'{c}_개수'] = n
            rows.append(row)
    return pd.DataFrame(rows)


def _apply_diff_formatting(ws, df, col_pairs):
    """df의 (A컬럼명, B컬럼명) 쌍마다 값이 다르면 두 셀 모두 배경색을 칠한다."""
    if df.empty:
        return
    headers = list(df.columns)

    def _to_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    pairs = [(headers.index(a) + 1, headers.index(b) + 1) for a, b in col_pairs if a in headers and b in headers]
    for row_idx in range(2, ws.max_row + 1):
        for col_a, col_b in pairs:
            val_a = ws.cell(row=row_idx, column=col_a).value
            val_b = ws.cell(row=row_idx, column=col_b).value
            if abs(_to_float(val_a) - _to_float(val_b)) > 1e-4:
                ws.cell(row=row_idx, column=col_a).fill = DIFF_FILL
                ws.cell(row=row_idx, column=col_b).fill = DIFF_FILL


def build_comparison_workbook(data_a, data_b, floor_mapping, spaces_dict_a, spaces_dict_b,
                               is_llm_floor_mapping=False, llm_floor_detail=None,
                               area_thresh=2.0, centroid_thresh=1.0, status_cb=None):
    """비교 결과 워크북을 생성해 반환한다(저장은 호출부에서)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    if status_cb:
        status_cb('층 매칭 시트 작성 중...')
    df_floor = build_floor_mapping_df(floor_mapping, is_llm=is_llm_floor_mapping, llm_detail=llm_floor_detail)
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '01_층_매칭', used_names))
    ite._write_df(ws, df_floor)

    matched_pairs, unmatched_a, unmatched_b, df_space = compute_all_space_matches(
        data_a, data_b, floor_mapping, spaces_dict_a, spaces_dict_b,
        area_thresh=area_thresh, centroid_thresh=centroid_thresh, status_cb=status_cb,
    )
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '02_공간_매칭', used_names))
    ite._write_df(ws, df_space)

    if status_cb:
        status_cb('구조부재 집계 계산 중...')
    breakdowns = _compute_breakdowns(data_a, data_b, matched_pairs, status_cb=status_cb)

    df_split = build_structural_split_df(breakdowns)
    ws_split = wb.create_sheet(ite._unique_sheet_name(wb, '03_구조부재_내외구분비교', used_names))
    ite._write_df(ws_split, df_split)
    _apply_diff_formatting(ws_split, df_split, [('A_개수', 'B_개수'), ('A_면적(㎡)', 'B_면적(㎡)')])

    df_total = build_structural_total_df(breakdowns)
    ws_total = wb.create_sheet(ite._unique_sheet_name(wb, '03b_구조부재_통합비교', used_names))
    ite._write_df(ws_total, df_total)
    _apply_diff_formatting(ws_total, df_total, [('A_개수', 'B_개수'), ('A_면적(㎡)', 'B_면적(㎡)')])

    if status_cb:
        status_cb('층단위 부재 비교 계산 중 (공간 매칭 무관, 저비용)...')
    df_floor_level = build_floor_level_comparison_df(data_a, data_b, floor_mapping)
    ws_floor_level = wb.create_sheet(ite._unique_sheet_name(wb, '03c_층단위_부재비교', used_names))
    ite._write_df(ws_floor_level, df_floor_level)
    _apply_diff_formatting(ws_floor_level, df_floor_level, [('A_개수', 'B_개수'), ('A_면적(㎡)', 'B_면적(㎡)')])

    if status_cb:
        status_cb('설비 비교 계산 중...')
    df_equip = build_equipment_comparison_df(data_a['ifc_file'], data_b['ifc_file'], matched_pairs)
    ws_equip = wb.create_sheet(ite._unique_sheet_name(wb, '04_설비_비교', used_names))
    ite._write_df(ws_equip, df_equip)
    _apply_diff_formatting(ws_equip, df_equip, [('A_개수', 'B_개수')])

    if status_cb:
        status_cb('미매칭 공간 객체 요약 계산 중...')
    df_unmatched = build_unmatched_object_summary_df(data_a['ifc_file'], unmatched_a, data_b['ifc_file'], unmatched_b)
    ws = wb.create_sheet(ite._unique_sheet_name(wb, '05_미매칭공간_객체요약', used_names))
    ite._write_df(ws, df_unmatched)

    return wb
