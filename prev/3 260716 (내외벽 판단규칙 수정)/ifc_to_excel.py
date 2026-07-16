# -*- coding: utf-8 -*-
"""
ifc_to_excel.py
----------------
임의의 IFC 파일(.ifc)을 입력받아, 건축공학의 일반적 정보 계층
(Project > Site > Building > Storey > Space > Element)에 맞춰
행렬(엔티티 x 속성) 구조의 엑셀 파일로 자동 추출하는 모듈.

사용법 (CLI):
    python ifc_to_excel.py 입력파일.ifc [출력파일.xlsx]

사용법 (함수 호출):
    from ifc_to_excel import extract_ifc_to_excel
    extract_ifc_to_excel("model.ifc", "model_추출.xlsx")
"""

import sys
import argparse
import json
import math
import re
from collections import Counter, OrderedDict, defaultdict

import ifcopenshell
import ifcopenshell.util.element as E
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------
# 클래스명 -> 한글 시트명 매핑 (없는 클래스는 영문 클래스명 그대로 사용)
# 필요시 자유롭게 추가/수정 가능
# -----------------------------------------------------------------
CATEGORY_LABELS = {
    'IfcSite': '대지(Site)',
    'IfcBuilding': '건물(Building)',
    'IfcBuildingStorey': '층(Storey)',
    'IfcSpace': '공간(Space)',
    'IfcWall': '벽(Wall)',
    'IfcWallStandardCase': '벽(Wall)',
    'IfcCurtainWall': '커튼월(CurtainWall)',
    'IfcColumn': '기둥(Column)',
    'IfcBeam': '보(Beam)',
    'IfcMember': '부재(Member)',
    'IfcSlab': '바닥(Slab)',
    'IfcCovering': '천장외장(Covering)',
    'IfcRoof': '지붕(Roof)',
    'IfcRailing': '난간(Railing)',
    'IfcStair': '계단(Stair)',
    'IfcStairFlight': '계단참(StairFlight)',
    'IfcRamp': '경사로(Ramp)',
    'IfcRampFlight': '경사로참(RampFlight)',
    'IfcDoor': '문(Door)',
    'IfcWindow': '창(Window)',
    'IfcOpeningElement': '개구부(Opening)',
    'IfcPlate': '플레이트(Plate)',
    'IfcGrid': '그리드(Grid)',
    'IfcFurnishingElement': '적재물(Furnishing)',
    'IfcFlowTerminal': '설비단말(FlowTerminal)',
    'IfcDistributionElement': '설비요소(Distribution)',
    'IfcLightFixture': '조명(LightFixture)',
    'IfcSensor': '센서(Sensor)',
    'IfcFireSuppressionTerminal': '소방장치(FireSuppressionTerminal)',
    'IfcAlarm': '경보기(Alarm)',
}

CORE_ATTRS = ['GlobalId', 'Name', 'Description', 'ObjectType', 'Tag', 'PredefinedType', 'LongName']

# 행렬(Wide) 시트를 만들지 않을 비-요소성 IfcProduct 하위클래스(필요시 조정)
EXCLUDE_FROM_WIDE = {'IfcAnnotation'}

MAX_ROWS_LONG = 300_000  # 과도하게 큰 IFC 보호용 상한


# ===================================================================
# 내부 유틸
# ===================================================================

def _safe_attr(ent, name):
    try:
        v = getattr(ent, name)
    except Exception:
        return None
    if v is None:
        return None
    if hasattr(v, 'is_a'):
        try:
            return v.is_a()
        except Exception:
            return str(v)
    return v


def _storey_via_decomposition(ent):
    """ContainsElements(공간적 포함관계) 대신 Decomposes(IfcRelAggregates, 분해/구성관계)로만
    상위에 연결된 경우의 fallback. IfcSpace 등에서 흔히 나타남."""
    cur = ent
    seen = set()
    while cur is not None and hasattr(cur, 'is_a'):
        if cur.is_a('IfcBuildingStorey'):
            return cur
        if id(cur) in seen:
            break
        seen.add(id(cur))
        decomposes = getattr(cur, 'Decomposes', None)
        if not decomposes:
            return None
        cur = decomposes[0].RelatingObject
    return None


def _get_storey(ent):
    try:
        c = E.get_container(ent)
    except Exception:
        c = None
    cur = c
    seen = set()
    while cur is not None and hasattr(cur, 'is_a') and not cur.is_a('IfcBuildingStorey'):
        if id(cur) in seen:
            break
        seen.add(id(cur))
        try:
            cur = E.get_container(cur)
        except Exception:
            cur = None
    if cur is None:
        # 공간적 포함관계(ContainsElements)로 못 찾으면 분해/구성관계(Decomposes)로 재시도
        cur = _storey_via_decomposition(ent)
    return cur


def _get_spaces_via_boundary(ifc_file, ent):
    spaces = _get_spaces_list_via_boundary(ifc_file, ent)
    return ', '.join(spaces) if spaces else None


def _get_spaces_list_via_boundary(ifc_file, ent):
    spaces = []
    for rel in ifc_file.by_type('IfcRelSpaceBoundary'):
        if rel.RelatedBuildingElement == ent and rel.RelatingSpace is not None:
            spaces.append(rel.RelatingSpace.Name)
    return spaces


def _get_material(ent):
    try:
        m = E.get_material(ent)
    except Exception:
        return None
    if m is None:
        return None
    try:
        if m.is_a('IfcMaterial'):
            return m.Name
        if m.is_a('IfcMaterialList'):
            return ', '.join([mm.Name for mm in m.Materials])
        if m.is_a() == 'IfcMaterialLayerSetUsage' and m.ForLayerSet:
            return m.ForLayerSet.LayerSetName or 'LayerSet'
        if hasattr(m, 'Name') and m.Name:
            return m.Name
        return m.is_a()
    except Exception:
        return None


def _flatten_psets(ent):
    try:
        psets = E.get_psets(ent, qtos_only=False)
    except Exception:
        psets = {}
    flat = {}
    for pset_name, props in psets.items():
        for k, v in props.items():
            if k == 'id':
                continue
            flat[f"{pset_name}.{k}"] = v
    return flat


# -------------------------------------------------------------
# 좌표 기반 직접 치수 계산 (geometry 커널 미사용, 순수 좌표 파싱)
# 우선순위: ① IfcBoundingBox(Box표현) 직접값 > ② Extrude 단면 좌표+Depth > ③ Tessellation 좌표 bbox
# -------------------------------------------------------------

def _resolve_body_items(ent):
    """Representation 중 'Body' 식별자의 Item들을 MappedItem까지 풀어서 반환"""
    if not ent.Representation:
        return []
    items = []
    for rep in ent.Representation.Representations:
        if rep.RepresentationIdentifier != 'Body':
            continue
        for it in rep.Items:
            if it.is_a('IfcMappedItem'):
                mr = it.MappingSource.MappedRepresentation
                items.extend(mr.Items)
            else:
                items.append(it)
    return items


def _profile_xy_extent(profile):
    """IfcProfileDef 윤곽선 좌표에서 X/Y 폭 직접 계산 (mm 단위, IFC 원본 단위 그대로)"""
    try:
        if profile.is_a('IfcRectangleProfileDef'):
            return profile.XDim, profile.YDim
        if profile.is_a('IfcCircleProfileDef'):
            return profile.Radius * 2, profile.Radius * 2
        curve = getattr(profile, 'OuterCurve', None)
        if curve is not None and curve.is_a('IfcIndexedPolyCurve'):
            pts = curve.Points
            if pts.is_a('IfcCartesianPointList2D'):
                arr = np.array(pts.CoordList, dtype=float)
                ext = arr.max(axis=0) - arr.min(axis=0)
                return float(ext[0]), float(ext[1])
    except Exception:
        pass
    return None, None


def _get_local_dimensions(ent):
    """(X, Y, Z, 산출방식) 반환. 단위는 IFC 원본 길이단위(보통 mm) 그대로,
    호출부에서 단위 변환."""
    # ① IfcBoundingBox 표현이 있으면 그 값을 그대로 사용 (가장 신뢰도 높음, 파일 내 기록값)
    if ent.Representation:
        for rep in ent.Representation.Representations:
            if rep.RepresentationIdentifier == 'Box':
                for it in rep.Items:
                    if it.is_a('IfcBoundingBox'):
                        return it.XDim, it.YDim, it.ZDim, 'BoundingBox(원본기록값)'

    items = _resolve_body_items(ent)

    # ② 압출 형상: 단면(SweptArea) 윤곽 좌표의 X/Y 폭 + Depth(돌출방향=Z) 직접 계산
    for it in items:
        if it.is_a('IfcExtrudedAreaSolid'):
            x, y = _profile_xy_extent(it.SweptArea)
            if x is not None:
                return x, y, it.Depth, '단면좌표+Depth(직접계산)'

    # ③ 테셀레이션(메쉬): 전체 정점 좌표의 최대-최소 폭 직접 계산
    for it in items:
        if it.is_a('IfcPolygonalFaceSet'):
            coords = it.Coordinates.CoordList
            arr = np.array(coords, dtype=float)
            ext = arr.max(axis=0) - arr.min(axis=0)
            return float(ext[0]), float(ext[1]), float(ext[2]), '메쉬좌표bbox(직접계산)'

    return None, None, None, None


def _dimension_columns(ent, length_unit_scale=0.001):
    """엑셀 컬럼용 치수 dict. length_unit_scale: IFC 원본단위(mm 가정) -> m 변환 계수"""
    x, y, z, src = _get_local_dimensions(ent)
    conv = lambda v: round(v * length_unit_scale, 4) if v is not None else None
    return {
        '치수_X(m)': conv(x),
        '치수_Y(m)': conv(y),
        '치수_Z(m)': conv(z),
        '치수산출방식': src,
    }


DIMENSION_TARGET_CLASSES = {'IfcWall', 'IfcWallStandardCase', 'IfcColumn', 'IfcBeam', 'IfcMember', 'IfcCurtainWall'}


# -------------------------------------------------------------
# 면적 직접 계산 (좌표 기반, 신발끈 공식 등) + Pset 값 폴백
# 우선순위: ① Extrude 단면(SweptArea) 실제 면적(신발끈공식/도형공식) 직접계산
#         > ② Tessellation(메쉬) 3D 삼각형 합산 면적 직접계산
#         > ③ 좌표 기반 계산이 불가능한 경우 Pset/Qto에 기재된 면적류 속성값 사용(폴백)
# -------------------------------------------------------------

def _polyline_points_2d(curve):
    """IfcIndexedPolyCurve 또는 IfcPolyline에서 2D 좌표 포인트 리스트 반환 (원본단위, 보통 mm)."""
    try:
        if curve.is_a('IfcIndexedPolyCurve'):
            pts = curve.Points
            if pts.is_a('IfcCartesianPointList2D'):
                return [tuple(p) for p in pts.CoordList]
        elif curve.is_a('IfcPolyline'):
            return [tuple(p.Coordinates) for p in curve.Points]
    except Exception:
        pass
    return None


def _shoelace_area(points):
    """신발끈 공식(Shoelace formula)으로 폐다각형 면적 계산. points: [(x,y), ...] 순서대로."""
    n = len(points)
    if n < 3:
        return None
    area = 0.0
    for i in range(n):
        x1, y1 = points[i][0], points[i][1]
        x2, y2 = points[(i + 1) % n][0], points[(i + 1) % n][1]
        area += x1 * y2 - x2 * y1
    return abs(area) / 2.0


def _profile_area(profile):
    """IfcProfileDef 실제 면적(원본단위^2, 보통 mm^2) 계산.
    사각형/원은 도형 공식, 임의 다각형(IfcArbitraryClosedProfileDef 등)은 신발끈 공식.
    IfcArbitraryProfileDefWithVoids는 내부 홀 면적을 빼준다."""
    try:
        if profile.is_a('IfcRectangleProfileDef'):
            return profile.XDim * profile.YDim
        if profile.is_a('IfcCircleProfileDef'):
            return math.pi * profile.Radius ** 2
        outer = getattr(profile, 'OuterCurve', None)
        if outer is None:
            return None
        pts = _polyline_points_2d(outer)
        if not pts:
            return None
        area = _shoelace_area(pts)
        if area is None:
            return None
        if profile.is_a('IfcArbitraryProfileDefWithVoids'):
            for inner in profile.InnerCurves:
                ipts = _polyline_points_2d(inner)
                if ipts:
                    iarea = _shoelace_area(ipts)
                    if iarea:
                        area -= iarea
        return area
    except Exception:
        return None


def _mesh_surface_area(coords, faces_idx):
    """정점좌표(coords, 0-based 배열)와 면 인덱스 목록(faces_idx, 각 면은 0-based 인덱스 튜플)으로
    3D 삼각형 합산 표면적 계산 (원본단위^2). 사각형 등 n각형 면은 삼각팬(triangle fan)으로 분할하며,
    경사면(지붕 등)도 실제 경사 표면적으로 정확히 계산된다(수평 투영 면적이 아님)."""
    arr = np.array(coords, dtype=float)
    total = 0.0
    for idx in faces_idx:
        if len(idx) < 3:
            continue
        p0 = arr[idx[0]]
        for k in range(1, len(idx) - 1):
            p1, p2 = arr[idx[k]], arr[idx[k + 1]]
            total += float(np.linalg.norm(np.cross(p1 - p0, p2 - p0))) / 2.0
    return total


def _get_footprint_area(ent):
    """(면적, 산출방식) 반환. 좌표 기반 직접계산만 시도하며, 실패시 (None, None)로
    호출부에서 Pset 폴백을 타도록 한다. 단위는 원본단위^2(보통 mm^2) 그대로."""
    items = _resolve_body_items(ent)

    # ① 압출 형상: 단면(SweptArea) 실제 면적 (신발끈공식/도형공식)
    for it in items:
        if it.is_a('IfcExtrudedAreaSolid'):
            area = _profile_area(it.SweptArea)
            if area is not None:
                return area, '단면좌표(신발끈공식) 직접계산'

    # ② 테셀레이션(메쉬): 3D 삼각형 합산 (경사면도 실제 표면적으로 계산됨)
    for it in items:
        if it.is_a('IfcPolygonalFaceSet'):
            coords = it.Coordinates.CoordList
            faces_idx = []
            for face in it.Faces:
                ci = face.CoordIndex
                faces_idx.append(tuple(i - 1 for i in ci))
            if faces_idx:
                return _mesh_surface_area(coords, faces_idx), '메쉬삼각형합산 직접계산'
        if it.is_a('IfcTriangulatedFaceSet'):
            coords = it.Coordinates.CoordList
            faces_idx = [tuple(i - 1 for i in tri) for tri in it.CoordIndex]
            if faces_idx:
                return _mesh_surface_area(coords, faces_idx), '메쉬삼각형합산 직접계산'

    return None, None


_AREA_KEY_RE = re.compile(r'area', re.IGNORECASE)
_RATIO_KEY_RE = re.compile(r'ratio', re.IGNORECASE)


def _get_qto_area(flat_props):
    """flat_props에서 '정식 Qto_*BaseQuantities' 면적류 속성만 찾는다(Pset_ 커스텀 속성은
    제외 - 그건 _get_pset_area_fallback이 최후 폴백으로 따로 처리한다).
    BIM 저작툴이 직접 계산해 내보낸 값이므로 우리 자체 지오메트리 계산(footprint/시스템
    bounding 등)보다 신뢰도가 높다고 보고 최우선으로 사용한다.
    우선순위: Gross > Net > 기타(단순 Area 등)."""
    candidates = []
    for key, val in flat_props.items():
        if not isinstance(val, (int, float)):
            continue
        if not key.startswith('Qto_'):
            continue
        if not _AREA_KEY_RE.search(key) or _RATIO_KEY_RE.search(key):
            continue
        score = 0 if 'Gross' in key else (1 if 'Net' in key else 2)
        candidates.append((score, key, val))
    if not candidates:
        return None, None
    candidates.sort(key=lambda c: c[0])
    _, key, val = candidates[0]
    return val, key


def _get_pset_area_fallback(flat_props):
    """flat_props({'Pset.속성': 값, ...})에서 면적류 속성을 찾아 (값, 출처키) 반환.
    'Ratio'가 포함된 속성(예: Reinforcement_Area_Ratio)은 면적이 아니라 제외.
    Qto_ 접두사(정식 수량세트)를 Pset_ 커스텀 속성보다 우선하고, 'Gross'가 포함된 것을 우선한다."""
    candidates = []
    for key, val in flat_props.items():
        if not isinstance(val, (int, float)):
            continue
        if not _AREA_KEY_RE.search(key) or _RATIO_KEY_RE.search(key):
            continue
        score = (0 if key.startswith('Qto_') else 1, 0 if 'Gross' in key else 1)
        candidates.append((score, key, val))
    if not candidates:
        return None, None
    candidates.sort(key=lambda c: c[0])
    _, key, val = candidates[0]
    return val, key


SYSTEM_ASSEMBLY_CLASSES = {'IfcDoor', 'IfcWindow', 'IfcCurtainWall'}
# 문/창/커튼월은 실무에서 "부품(멀리언/프레임/패널/하드웨어) 개별 면적"이 아니라
# "그 시스템 전체의 대표 면적(개구부/입면 면적)"에 관심이 있다는 요구사항에 따른 클래스 목록.
# 이런 부재는 흔히 프레임/하드웨어 디테일이 섞인 상세 메쉬로 모델링되어, 메쉬 전체
# 표면적을 합산하는 _get_footprint_area()가 실제 면적보다 훨씬 큰 값을 준다(실측:
# 문 하나가 15㎡로 나온 사례 - 프레임/패널 등 여러 면이 겹쳐 합산됨).


def _get_system_bounding_face_area(ent):
    """문/창/커튼월처럼 하위부품으로 구성된 '시스템' 부재의 전체 대표 면적을 계산한다.
    부품 개별 면적이 아니라 전체 bounding 치수(_get_local_dimensions, 이미 실측 검증된
    안정적 함수) 중 가장 큰 두 축(=폭 x 높이)을 곱하고, 가장 작은 축(=두께)은 제외한다.
    반환: (면적, 산출방식) 또는 계산 불가시 (None, None). 단위는 원본단위^2 그대로."""
    x, y, z, src = _get_local_dimensions(ent)
    dims = [d for d in (x, y, z) if d is not None]
    if len(dims) < 2:
        return None, None
    dims.sort(reverse=True)
    area = dims[0] * dims[1]
    return area, f'시스템 전체 bounding치수 기반(폭x높이, 두께제외; {src})'


def _get_decomposed_children(ent, _depth=0, _max_depth=3):
    """IsDecomposedBy(IfcRelAggregates)로 연결된 하위부품들을 재귀적으로 수집.
    _max_depth로 무한재귀/과도한 깊이를 방지(일반적으로 커튼월->멀리언/패널 1단계면 충분)."""
    if _depth >= _max_depth:
        return []
    children = []
    for rel in (getattr(ent, 'IsDecomposedBy', None) or []):
        for child in rel.RelatedObjects:
            children.append(child)
            children.extend(_get_decomposed_children(child, _depth + 1, _max_depth))
    return children


def _get_assembly_bounding_face_area(ent):
    """IfcCurtainWall처럼 부재 자신은 직접 지오메트리(Representation)가 없고, 하위부품
    (멀리언=IfcMember/패널=IfcPlate 등, IsDecomposedBy)에만 지오메트리가 있는 '조립체
    컨테이너'의 전체 대표 면적을 계산한다(실측으로 이런 구조를 확인함: 커튼월 자체는
    Representation이 없고 자식으로 IfcMember/IfcPlate만 있음).
    각 하위부품의 로컬 bounding 치수 + 월드 배치행렬로 bounding box 8개 모서리를 월드좌표로
    변환해 전부 모으고, 전체를 감싸는 bounding box에서 가장 큰 두 축(폭x높이, 두께 제외)을
    면적으로 쓴다. ifcopenshell.geom의 지오메트리 커널(삼각분할)은 쓰지 않고 배치행렬만
    사용해 가볍다 - 다만 부재 로컬원점이 bounding box 코너와 정확히 일치한다는 보장은 없어
    "전체 시스템이 대략 이 정도 규모"라는 근사치임을 산출방식 문구에 명시한다.
    반환: (면적, 산출방식) 또는 (None, None)."""
    import ifcopenshell.util.placement as plc

    children = _get_decomposed_children(ent)
    if not children:
        return None, None

    all_corners = []
    for child in children:
        x, y, z, _src = _get_local_dimensions(child)
        if x is None or y is None or z is None:
            continue
        if child.ObjectPlacement is None:
            continue
        try:
            m = plc.get_local_placement(child.ObjectPlacement)
        except Exception:
            continue
        for sx in (0, x):
            for sy in (0, y):
                for sz in (0, z):
                    world_pt = m @ np.array([sx, sy, sz, 1.0])
                    all_corners.append(world_pt[:3])

    if len(all_corners) < 2:
        return None, None
    arr = np.array(all_corners)
    extent = arr.max(axis=0) - arr.min(axis=0)
    dims = sorted(extent, reverse=True)
    area = float(dims[0] * dims[1])
    return area, f'하위부품{len(children)}개 월드bounding 근사(폭x높이, 두께제외)'


def _area_columns(ent, flat_props, length_unit_scale=0.001):
    """엑셀 컬럼용 면적 dict. 우선순위(중요 - 사용자 요청에 따른 순서):
    ⓪ Qto_*BaseQuantities 정식 수량값(_get_qto_area) - BIM 저작툴이 직접 계산해 낸
       값이라 최우선으로 신뢰한다(우리 자체 계산보다 먼저 시도).
    ① (⓪이 없을 때만) IfcDoor/IfcWindow/IfcCurtainWall(SYSTEM_ASSEMBLY_CLASSES)은 부품
       단위 메쉬 합산 대신 전체 bounding 치수(_get_system_bounding_face_area, 부재
       자신에 지오메트리가 있는 경우), 그것도 없으면 하위부품 전체를 감싸는 bounding
       (_get_assembly_bounding_face_area, 커튼월처럼 멀리언/패널에만 지오메트리가
       있는 경우)을 사용한다 - 부품 단위 메쉬 합산은 실제보다 훨씬 큰 값을 주는 경우가
       실측으로 확인되었기 때문.
    ② 그 외 클래스 및 ①에서 실패한 경우: 좌표 직접계산(footprint/메쉬).
    ③ ①②모두 실패시: Qto가 아닌 Pset 커스텀 속성값 폴백(_get_pset_area_fallback).
    length_unit_scale: 원본단위->m 변환 계수(①②는 좌표기반이라 적용, ⓪③은 Pset/Qto
    값 자체가 이미 프로젝트 단위(보통 m²)로 저장되어 있어 그대로 사용, 추가 변환 없음)."""
    area, source_key = _get_qto_area(flat_props)
    if area is not None:
        return {'면적(㎡)': round(area, 4), '면적산출방식': f'Qto값 사용({source_key})'}

    area, method = None, None
    if ent.is_a() in SYSTEM_ASSEMBLY_CLASSES:
        area, method = _get_system_bounding_face_area(ent)
        if area is None:
            area, method = _get_assembly_bounding_face_area(ent)

    if area is None:
        area, method = _get_footprint_area(ent)

    if area is None:
        area, source_key = _get_pset_area_fallback(flat_props)
        if area is not None:
            method = f'Pset값 사용({source_key})'
        else:
            method = '산출불가(좌표계산 실패 + Pset값 없음)'
    else:
        area = area * (length_unit_scale ** 2)
    return {
        '면적(㎡)': round(area, 4) if area is not None else None,
        '면적산출방식': method,
    }


AREA_TARGET_CLASSES = {'IfcRoof', 'IfcCovering', 'IfcSlab', 'IfcDoor', 'IfcWindow', 'IfcCurtainWall'}



# ===================================================================
# 3-1b. 벽 내/외벽 판정 (2단계)
#   1차: Pset_WallCommon.IsExternal 속성값 (모델러가 직접 기입한 값 -> 있으면 최우선 신뢰)
#   2차(1차 값이 없을 때만): RelSpaceBoundary 경계 지오메트리로 벽 양쪽 면에 각각 Space가
#       존재하는지 확인. 양쪽 다 확인되면 '내벽'. 한쪽만 확인되면 반대면은 Space 경계로
#       기록되지 않았다는 뜻이므로 '외벽(추정)'으로 분류한다 (Pset처럼 명시적으로 기입된
#       값이 아니라 간접 추정이라는 점을 라벨에서 구분 - 오류가 아니라 정상적인 2차 추정 결과).
# ===================================================================

def _wall_pset_is_external(ent):
    """Pset_WallCommon.IsExternal 값을 bool로 반환. 없거나 bool이 아니면 None."""
    try:
        psets = E.get_psets(ent, psets_only=True)
    except Exception:
        return None
    val = (psets.get('Pset_WallCommon') or {}).get('IsExternal')
    return val if isinstance(val, bool) else None


def _wall_both_sides_space_check(ifc_file):
    """RelSpaceBoundary.ConnectionGeometry를 world 좌표로 변환해, 벽 GlobalId별로
    서로 반대 방향(법선) 양쪽에 각각 Space 경계가 존재하는지 확인.
    반환: {GlobalId: True(양쪽 확인)/False(한쪽만 확인)}. 경계 지오메트리가 아예 없는
    벽은 키 자체가 없음(→ 호출부에서 '판정불가'로 처리)."""
    import ifcopenshell.util.placement as plc

    wall_normals = {}
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        if elem is None or not elem.is_a('IfcWall'):
            continue
        space = r.RelatingSpace
        cg = r.ConnectionGeometry
        if space is None or cg is None:
            continue
        surf = cg.SurfaceOnRelatingElement
        if surf is None or not surf.is_a('IfcCurveBoundedPlane'):
            continue
        plane = surf.BasisSurface
        try:
            space_m = plc.get_local_placement(space.ObjectPlacement)
            plane_m = plc.get_axis2placement(plane.Position)
        except Exception:
            continue
        world_m = space_m @ plane_m
        normal = world_m[:3, 2]
        nrm = np.linalg.norm(normal)
        if nrm == 0:
            continue
        wall_normals.setdefault(elem.GlobalId, []).append(normal / nrm)

    result = {}
    for gid, normals in wall_normals.items():
        ref = normals[0]
        has_same = any(np.dot(n, ref) > 0.5 for n in normals)
        has_opp = any(np.dot(n, ref) < -0.5 for n in normals)
        result[gid] = has_same and has_opp
    return result


def _wall_distinct_space_count(ifc_file):
    """벽 GlobalId -> RelSpaceBoundary 관계로 연결된 서로 다른 Space GlobalId 개수.
    ConnectionGeometry 유무와 무관하게 '관계'만 본다 (3차 판정용, 2차보다 근거가 약함).
    주의(알려진 한계): 벽 하나의 같은 면이 여러 방을 순서대로 접해도 관계상 여러 Space로
    잡히므로, 이 개수만으로는 '진짜 양쪽 분리'와 '한쪽 면이 여러 방과 연속 접함'을 완벽히
    구분하지 못한다. ConnectionGeometry가 아예 없어 2차 판정 자체가 불가능한 경우에만
    판정불가보다 나은 참고 정보로 사용한다."""
    counts = defaultdict(set)
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        if elem is None or not elem.is_a('IfcWall') or r.RelatingSpace is None:
            continue
        counts[elem.GlobalId].add(r.RelatingSpace.GlobalId)
    return {gid: len(spaces) for gid, spaces in counts.items()}


def _determine_wall_classification(ifc_file):
    """벽 GlobalId -> ('내벽'/'외벽'/'외벽(추정)'/'내벽(추정-관계기반)'/'판정불가', 판정근거) 딕셔너리.
    1차: Pset_WallCommon.IsExternal.
        ** 2026-07 기준 비활성화(주석처리) ** - 불완전하게 export된 IFC에서도 이 속성
        자체는 존재하지만 실제 값(True/False)이 사실과 다르게 채워지는 경우가 확인되어,
        이 값을 신뢰할 수 없다고 판단해 1차 판정을 건너뛰도록 했다. 코드는 나중에 다시
        쓸 수 있도록 삭제하지 않고 주석처리만 해둔다 - 재활성화하려면 아래
        "# is_ext = _wall_pset_is_external(w)"와 그 아래 두 if/elif 블록의 주석을 해제하면 된다.
    2차(1차 없을 때만 - 지금은 항상): RelSpaceBoundary.ConnectionGeometry 기반 양면 Space 접촉 확인.
    3차(1차/2차 모두 근거 없을 때만, 예: ConnectionGeometry 자체가 파일에 없는 경우):
        지오메트리 없이 RelSpaceBoundary '관계'만으로 서로 다른 Space에 2개 이상 연결되는지
        확인. 2차보다 근거가 약해 '추정-관계기반'으로 명확히 구분 표시한다."""
    both_sides = _wall_both_sides_space_check(ifc_file)
    distinct_space_count = _wall_distinct_space_count(ifc_file)
    result = {}
    for w in ifc_file.by_type('IfcWall'):
        gid = w.GlobalId
        # 1차(Pset_WallCommon.IsExternal) 판정 비활성화 - 위 docstring 참고.
        # is_ext = _wall_pset_is_external(w)
        is_ext = None
        if is_ext is True:
            result[gid] = ('외벽', '1차: Pset_WallCommon.IsExternal=True')
        elif is_ext is False:
            result[gid] = ('내벽', '1차: Pset_WallCommon.IsExternal=False')
        elif gid in both_sides:
            if both_sides[gid]:
                result[gid] = ('내벽', '2차: RelSpaceBoundary 양면 Space 접촉 확인')
            else:
                result[gid] = ('외벽(추정)', '2차: 한쪽 면만 Space 접촉 → 외벽으로 추정(반대면 Space 경계 없음)')
        elif distinct_space_count.get(gid, 0) >= 2:
            n = distinct_space_count[gid]
            result[gid] = ('내벽(추정-관계기반)',
                           f'3차: ConnectionGeometry 없음, RelSpaceBoundary 관계상 서로 다른 Space {n}개와 '
                           f'연결됨(지오메트리 미확인 - 같은 면이 여러 방과 연속 접한 경우일 수도 있어 추정치임)')
        else:
            result[gid] = ('판정불가', '1차/2차/3차 모두 근거 데이터 없음')
    return result


# ===================================================================
# 3-1c. 벽 외 모든 구조부재 대상 범용 내/외부 판정 (위 벽 전용 로직의 일반화 버전)
#   기존 _determine_wall_classification()은 하위호환을 위해 그대로 두고(라벨이 '내벽'/
#   '외벽'으로 벽 전용 문구), 여기서는 클래스에 무관하게 동작하며 라벨도 '내부'/'외부'로
#   중립적으로 표기한다. Pset_WallCommon 대신 "이름이 Common으로 끝나는 모든 Pset의
#   IsExternal"을 봐서 클래스별 Pset 이름(Pset_SlabCommon, Pset_DoorCommon 등)을
#   일일이 나열하지 않아도 되게 했다.
# ===================================================================

ELEMENT_CLASSIFICATION_TARGET_CLASSES = (
    'IfcWall', 'IfcWallStandardCase', 'IfcSlab', 'IfcRoof', 'IfcCovering',
    'IfcColumn', 'IfcBeam', 'IfcMember', 'IfcCurtainWall', 'IfcDoor', 'IfcWindow',
    'IfcRailing', 'IfcStair', 'IfcStairFlight', 'IfcRamp', 'IfcRampFlight',
)


def _element_pset_is_external(ent):
    """<임의Pset이름>Common.IsExternal 값을 bool로 반환 (클래스 무관 범용).
    _wall_pset_is_external(Pset_WallCommon 고정)의 일반화 버전."""
    try:
        psets = E.get_psets(ent, psets_only=True)
    except Exception:
        return None
    for pset_name, props in psets.items():
        if pset_name.endswith('Common'):
            val = props.get('IsExternal')
            if isinstance(val, bool):
                return val
    return None


def _element_both_sides_space_check(ifc_file, target_classes=None):
    """_wall_both_sides_space_check의 일반화 버전: RelatedBuildingElement의 클래스가
    target_classes(None이면 전체)에 속하는 모든 부재에 대해 양면 Space 접촉을 확인."""
    import ifcopenshell.util.placement as plc

    normals = {}
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        if elem is None:
            continue
        if target_classes is not None and elem.is_a() not in target_classes:
            continue
        space = r.RelatingSpace
        cg = r.ConnectionGeometry
        if space is None or cg is None:
            continue
        surf = cg.SurfaceOnRelatingElement
        if surf is None or not surf.is_a('IfcCurveBoundedPlane'):
            continue
        plane = surf.BasisSurface
        try:
            space_m = plc.get_local_placement(space.ObjectPlacement)
            plane_m = plc.get_axis2placement(plane.Position)
        except Exception:
            continue
        world_m = space_m @ plane_m
        normal = world_m[:3, 2]
        nrm = np.linalg.norm(normal)
        if nrm == 0:
            continue
        normals.setdefault(elem.GlobalId, []).append(normal / nrm)

    result = {}
    for gid, ns in normals.items():
        ref = ns[0]
        has_same = any(np.dot(n, ref) > 0.5 for n in ns)
        has_opp = any(np.dot(n, ref) < -0.5 for n in ns)
        result[gid] = has_same and has_opp
    return result


def _element_distinct_space_count(ifc_file, target_classes=None):
    """_wall_distinct_space_count의 일반화 버전. 알려진 한계(같은 면이 여러 방과 연속
    접하는 경우와 진짜 양쪽분리를 완전히 구분 못함)는 벽과 동일하게 적용된다."""
    counts = defaultdict(set)
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        if elem is None or r.RelatingSpace is None:
            continue
        if target_classes is not None and elem.is_a() not in target_classes:
            continue
        counts[elem.GlobalId].add(r.RelatingSpace.GlobalId)
    return {gid: len(spaces) for gid, spaces in counts.items()}


def _determine_element_classification(ifc_file, target_classes=ELEMENT_CLASSIFICATION_TARGET_CLASSES):
    """_determine_wall_classification의 일반화 버전 - 벽 뿐 아니라 target_classes에 속하는
    모든 구조부재에 대해 동일한 3단계 로직(Pset -> 양면Space접촉 -> 관계개수)으로 내/외부를
    판정한다. 반환: {GlobalId: (라벨, 판정근거)}, 라벨은 '내부'/'외부'/'외부(추정)'/
    '내부(추정-관계기반)'/'판정불가' (벽 전용 함수의 '내벽'/'외벽' 문구 대신 중립 표기).

    1차(Pset_*Common.IsExternal) 판정은 ** 2026-07 기준 비활성화(주석처리) ** - 벽과 동일한
    이유(_determine_wall_classification 참고: 불완전한 IFC에서도 속성은 존재하지만 값 자체를
    신뢰할 수 없음)로 건너뛴다. 재활성화하려면 아래 "# is_ext = _element_pset_is_external(ent)"와
    그 아래 두 if/elif 블록의 주석을 해제하면 된다."""
    both_sides = _element_both_sides_space_check(ifc_file, target_classes)
    distinct_count = _element_distinct_space_count(ifc_file, target_classes)

    result = {}
    for cls in target_classes:
        for ent in ifc_file.by_type(cls):
            gid = ent.GlobalId
            if gid in result:
                continue  # IfcWall/IfcWallStandardCase 등 상위-하위 클래스 중복 조회 방지
            # 1차(Pset_*Common.IsExternal) 판정 비활성화 - 위 docstring 참고.
            # is_ext = _element_pset_is_external(ent)
            is_ext = None
            if is_ext is True:
                result[gid] = ('외부', '1차: Pset_*Common.IsExternal=True')
            elif is_ext is False:
                result[gid] = ('내부', '1차: Pset_*Common.IsExternal=False')
            elif gid in both_sides:
                if both_sides[gid]:
                    result[gid] = ('내부', '2차: RelSpaceBoundary 양면 Space 접촉 확인')
                else:
                    result[gid] = ('외부(추정)', '2차: 한쪽 면만 Space 접촉 → 외부로 추정')
            elif distinct_count.get(gid, 0) >= 2:
                n = distinct_count[gid]
                result[gid] = ('내부(추정-관계기반)',
                               f'3차: ConnectionGeometry 없음, 서로 다른 Space {n}개와 연결(추정치)')
            else:
                result[gid] = ('판정불가', '1차/2차/3차 모두 근거 데이터 없음')
    return result


# ===================================================================
# 3-2. 공간(Space) - 부재(Element) 1:1 매칭 시트
#      IfcRelSpaceBoundary 1건 = 1행. 부재 1개가 여러 공간에 접하면
#      그만큼 행이 분리되어 나타남 (콤마로 뭉치지 않음)
# ===================================================================

def _build_space_element_matrix(ifc_file, wall_classification=None):
    if wall_classification is None:
        wall_classification = _determine_wall_classification(ifc_file)
    rows = []
    for rel in ifc_file.by_type('IfcRelSpaceBoundary'):
        sp = rel.RelatingSpace
        elem = rel.RelatedBuildingElement
        if sp is None or elem is None:
            continue
        storey = _storey_via_decomposition(sp) or _get_storey(sp)
        cls_result, cls_reason = (wall_classification.get(elem.GlobalId, (None, None))
                                   if elem.is_a('IfcWall') else (None, None))
        rows.append({
            '층(Storey)': storey.Name if storey else None,
            '공간(Space)_Name': sp.Name,
            '공간(Space)_GUID': sp.GlobalId,
            '부재_IFC_Class': elem.is_a(),
            '부재_Name': _safe_attr(elem, 'Name'),
            '부재_GUID': elem.GlobalId,
            'PhysicalOrVirtual': rel.PhysicalOrVirtualBoundary,
            'InternalOrExternal': rel.InternalOrExternalBoundary,
            '벽_내외벽_판정': cls_result,
            '벽_판정근거': cls_reason,
            'RelSpaceBoundary_GUID': rel.GlobalId,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['층(Storey)', '공간(Space)_Name', '부재_IFC_Class', '부재_Name']).reset_index(drop=True)
    return df


# ===================================================================
# 3-2b. 벽 내/외벽 판정 - 벽 1개당 1행 요약 시트
# ===================================================================

def _build_wall_classification_sheet(ifc_file, wall_classification=None):
    if wall_classification is None:
        wall_classification = _determine_wall_classification(ifc_file)
    rows = []
    for w in ifc_file.by_type('IfcWall'):
        storey = _get_storey(w)
        result, reason = wall_classification.get(w.GlobalId, ('판정불가', '근거 데이터 없음'))
        rows.append({
            '층(Storey)': storey.Name if storey else None,
            'Name': _safe_attr(w, 'Name'),
            'GlobalId': w.GlobalId,
            '내외벽_판정': result,
            '판정근거': reason,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['층(Storey)', '내외벽_판정', 'Name']).reset_index(drop=True)
    return df


# ===================================================================
# 3-2c. IFC4 명명규칙(표준 의도) 진단 - NVIDIA NIM LLM 이용 (선택 기능)
# ===================================================================
# 주의(사실 확인): 이 진단은 IFC4 스키마 자체를 코드로 검증하는 것이 아니라, LLM이 학습한
# IFC4 지식을 바탕으로 (IFC_Class, Name) 조합이 표준 의도와 명백히 모순되는지 "추론"한
# 것이다. 확정된 오류가 아니라 검토 참고용 의심 목록이며, 사내 코드성 명명 규칙(예:
# 'W6_200')은 표준에 규정된 바가 없어 정상으로 처리하도록 프롬프트에 명시했다.
# 비용 절감을 위해 부재 인스턴스 전체가 아니라 (IFC_Class, Name) "고유 조합" 단위로만
# LLM을 호출한다(동일 조합의 인스턴스가 100개여도 호출은 1회).

IFC4_NAMING_DIAGNOSIS_PROMPT = """당신은 IFC4(buildingSMART 표준) 스키마에 정통한 BIM 검토 전문가입니다.
아래는 어떤 IFC 모델에서 추출한 (IFC 엔티티 클래스, Name 속성) 조합 목록입니다.
각 조합에 대해 Name이 IFC4 표준이 의도하는 해당 엔티티 클래스의 의미
(예: IfcWall=벽, IfcSlab=바닥/슬래브, IfcColumn=기둥, IfcBeam=보, IfcDoor=문, IfcWindow=창 등)와
명백히 모순되는지만 판단하세요.

판정 기준:
- 사내 코드/도면 표기 관행(예: 'W6_200', 'RW4_400', 'C1', 'B12' 같은 임의 기호)은 표준에 규정된 바가
  없으므로 그 자체로는 '정상'으로 처리하세요.
- Name이 명백히 다른 부재 종류를 가리키는 경우만(예: IfcWall인데 Name에 '바닥'/'Slab'/'지붕' 등이 포함,
  혹은 IfcDoor인데 Name에 '창'/'Window'가 포함) '의심'으로 표시하세요.
- 판단 근거가 부족하면 '판단보류'로 표시하세요.

조합 목록:
{combo_list}

반드시 아래 JSON 배열 형식으로만 답하라. 다른 텍스트나 코드블록 표시(```)를 포함하지 마라.
[{{"class": "<IFC_Class>", "name": "<Name>", "판정": "정상|의심|판단보류", "사유": "<한 줄 근거>"}}, ...]
"""


def make_nim_llm_call(api_key, model="meta/llama-3.1-70b-instruct"):
    """NVIDIA NIM(OpenAI 호환) 엔드포인트 호출 함수. 이 기능을 쓰지 않으면 openai/httpx는
    임포트되지 않으므로(지연 임포트), 안 쓰는 사용자는 별도 설치가 필요 없다.
    사내 프록시 때문에 TLS 검증을 꺼야 하는 특수한 경우가 아니라면 verify=False는 쓰지 말 것
    (중간자 공격에 취약해짐). 꼭 꺼야 한다면 NIM_TLS_VERIFY=false 환경변수로만 명시적으로 끄도록 함
    (ifc_compare_core.py의 make_nim_llm_call과 동일한 방식)."""
    import os
    import httpx
    from openai import OpenAI
    tls_verify = os.environ.get("NIM_TLS_VERIFY", "true").lower() != "false"
    client = OpenAI(base_url="https://integrate.api.nvidia.com/v1", api_key=api_key,
                     http_client=httpx.Client(verify=tls_verify))

    def _call(prompt: str) -> str:
        resp = client.chat.completions.create(
            model=model, temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        return resp.choices[0].message.content
    return _call


def _collect_class_name_combos(ifc_file, target_classes=None):
    """(IFC_Class, Name) -> {count, sample_guid} 고유 조합 집계."""
    combos = OrderedDict()
    for e in ifc_file.by_type('IfcElement'):
        cls = e.is_a()
        if target_classes and cls not in target_classes:
            continue
        name = _safe_attr(e, 'Name') or '(이름없음)'
        key = (cls, name)
        if key not in combos:
            combos[key] = {'count': 0, 'sample_guid': e.GlobalId}
        combos[key]['count'] += 1
    return combos


def _parse_llm_json(raw):
    cleaned = raw.strip()
    if cleaned.startswith('```'):
        cleaned = cleaned.strip('`')
        if cleaned.lower().startswith('json'):
            cleaned = cleaned[4:]
    return json.loads(cleaned)


def run_ifc4_naming_diagnosis(ifc_file, llm_call, target_classes=None, batch_size=40, status_cb=None):
    """(IFC_Class, Name) 고유 조합별로 LLM에 IFC4 표준 의도 부합 여부를 물어 진단 DataFrame 반환.
    status_cb: str(message) -> None (선택, 진행상황 콜백)."""
    combos = _collect_class_name_combos(ifc_file, target_classes)
    keys = list(combos.keys())
    results = {}

    if status_cb:
        status_cb(f"IFC4 명명규칙 진단 시작: 고유 (클래스,이름) 조합 {len(keys)}개, "
                   f"배치크기 {batch_size} -> 약 {(len(keys) + batch_size - 1) // batch_size}회 LLM 호출 예정")

    for i in range(0, len(keys), batch_size):
        chunk = keys[i:i + batch_size]
        combo_list = "\n".join(f"- class: {c}, name: {n}" for c, n in chunk)
        prompt = IFC4_NAMING_DIAGNOSIS_PROMPT.format(combo_list=combo_list)
        try:
            raw = llm_call(prompt)
            parsed = _parse_llm_json(raw)
            for item in parsed:
                key = (item.get('class'), item.get('name'))
                results[key] = (item.get('판정', '판단보류'), item.get('사유', ''))
        except Exception as e:
            for key in chunk:
                results[key] = ('판단불가', f'LLM 호출/응답 파싱 실패: {e}')
            if status_cb and 'connection' in str(e).lower() and i == 0:
                status_cb("[힌트] 'Connection error'는 대부분 사내망/프록시가 NVIDIA API로의 "
                           "TLS 연결을 가로막는 경우입니다. NIM_TLS_VERIFY=false 환경변수를 설정해보세요 "
                           "(사내 프록시가 자체 인증서로 TLS를 가로채는 경우에만 사용 - 보안 트레이드오프 있음). "
                           "그래도 안 되면 사내 방화벽에서 integrate.api.nvidia.com(443)이 막혀있는지 "
                           "네트워크팀에 확인이 필요합니다.")
        if status_cb:
            status_cb(f"IFC4 명명규칙 진단: {min(i + batch_size, len(keys))}/{len(keys)} 조합 처리 완료")

    rows = []
    for (cls, name), info in combos.items():
        판정, 사유 = results.get((cls, name), ('판단불가', 'LLM 처리 안됨'))
        rows.append({
            'IFC_Class': cls, 'Name': name,
            '인스턴스_개수': info['count'], '샘플_GlobalId': info['sample_guid'],
            '판정(LLM추론)': 판정, '판정근거': 사유,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        order = {'의심': 0, '판단보류': 1, '판단불가': 2, '정상': 3}
        df['_sort'] = df['판정(LLM추론)'].map(order).fillna(9)
        df = df.sort_values(['_sort', 'IFC_Class', 'Name']).drop(columns='_sort').reset_index(drop=True)
    return df


# ===================================================================
# 3-3. 공간(Space)에 매칭되지 않은 부재 시트
#      사유를 자동분류: ①구조상 비대상 클래스 / ②해당층에 Space없음 / ③매칭누락의심
# ===================================================================

# ===================================================================
# 3-2b. RelSpaceBoundary 중복 진단
#   실측으로 확인된 사실: 같은 부재(벽/기둥/문/창 등) 하나가 하나의 Space와 여러 개의
#   별도 경계면(RelSpaceBoundary 레코드)으로 연결되는 경우가 있어, "관계 건수"를 그대로
#   개수로 쓰면 실제 물리적 개체 수보다 과다 집계된다 (예: 표본 파일에서 IfcWall은
#   744건 관계 중 실제 고유 개체는 438개 - 초과 306건, IfcColumn도 329건 중 82개 - 초과 247건).
#   04_공간-부재_매칭(1대1) 시트는 의도적으로 "관계 1건 = 1행"으로 그대로 보여주므로
#   (경계 데이터 자체를 살펴보기 위한 상세 시트), 그 시트의 행 수를 그대로 개수로 읽으면
#   안 된다는 것을 이 진단 시트로 함께 확인할 수 있게 한다.
# ===================================================================

def _build_relspaceboundary_duplication_report(ifc_file):
    """(Space, 부재) 쌍 기준으로, 같은 부재가 '같은 공간'에 RelSpaceBoundary로 여러 번
    중복 연결되는 경우를 클래스별로 집계한다.
    주의: 한 부재가 서로 다른 여러 공간에 연결되는 것은 정상이며 중복이 아니다(예: 방 2개에
    접한 벽은 관계가 2건 있는 게 맞음). 여기서는 '같은 (공간, 부재) 쌍'이 몇 번 반복되는지만
    보며, 이게 1보다 크면 get_space_related_elements()가 dedup해야 하는 진짜 중복이다."""
    pair_count = defaultdict(int)
    elem_class = {}
    for r in ifc_file.by_type('IfcRelSpaceBoundary'):
        elem = r.RelatedBuildingElement
        sp = r.RelatingSpace
        if elem is None or sp is None:
            continue
        pair_count[(sp.GlobalId, elem.GlobalId)] += 1
        elem_class[elem.GlobalId] = elem.is_a()

    raw_by_class = defaultdict(int)
    unique_pairs_by_class = defaultdict(int)
    for (sp_guid, elem_guid), cnt in pair_count.items():
        cls = elem_class[elem_guid]
        raw_by_class[cls] += cnt
        unique_pairs_by_class[cls] += 1

    rows = []
    for cls in sorted(set(raw_by_class) | set(unique_pairs_by_class)):
        raw = raw_by_class[cls]
        uniq = unique_pairs_by_class[cls]
        rows.append({
            'IFC_Class': cls,
            '원시_관계건수': raw,
            '고유_공간-부재쌍_수': uniq,
            '초과분(같은공간에중복연결)': raw - uniq,
            '중복여부': '있음' if raw > uniq else '없음',
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values('초과분(같은공간에중복연결)', ascending=False).reset_index(drop=True)
    return df


def _build_unmatched_elements(ifc_file):
    rsb = ifc_file.by_type('IfcRelSpaceBoundary')
    matched_guids = {rel.RelatedBuildingElement.GlobalId
                      for rel in rsb if rel.RelatedBuildingElement is not None}
    matched_classes = {rel.RelatedBuildingElement.is_a()
                        for rel in rsb if rel.RelatedBuildingElement is not None}

    space_storeys = set()
    for sp in ifc_file.by_type('IfcSpace'):
        st = _storey_via_decomposition(sp) or _get_storey(sp)
        if st:
            space_storeys.add(st.Name)

    rows = []
    for ent in ifc_file.by_type('IfcElement'):
        if ent.GlobalId in matched_guids:
            continue
        cls = ent.is_a()
        storey = _get_storey(ent)
        storey_name = storey.Name if storey else None

        if cls not in matched_classes:
            reason = '①구조상 비대상(이 모델에서 해당 클래스는 RelSpaceBoundary 미사용)'
        elif storey_name not in space_storeys:
            reason = '②해당 층에 Space 없음'
        else:
            reason = '③매칭누락 의심(같은 층에 Space 있는데 비어있음)'

        rows.append({
            '층(Storey)': storey_name,
            'IFC_Class': cls,
            'Name': _safe_attr(ent, 'Name'),
            'GlobalId': ent.GlobalId,
            'PredefinedType': _safe_attr(ent, 'PredefinedType'),
            '미매칭_사유': reason,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(['미매칭_사유', '층(Storey)', 'IFC_Class', 'Name']).reset_index(drop=True)
    return df


def _label(cls):
    return CATEGORY_LABELS.get(cls, cls)


# ===================================================================
# 1. 계층 구조 추출 (Project > Site > Building > Storey > Space)
# ===================================================================

def _build_hierarchy(ifc_file):
    rows = []

    def add(level, ifc_class, guid, name, parent_path, extra=''):
        rows.append({
            'Level': level, 'IFC_Class': ifc_class, 'GlobalId': guid,
            'Name': name, '상위경로': parent_path, '비고': extra,
        })

    projects = ifc_file.by_type('IfcProject')
    proj = projects[0] if projects else None
    if proj:
        add(0, 'IfcProject', proj.GlobalId, proj.Name, '')
    proj_name = proj.Name if proj else ''

    sites = ifc_file.by_type('IfcSite')
    if not sites:
        sites = [None]

    for site in sites:
        if site is None:
            buildings = ifc_file.by_type('IfcBuilding')
            path0 = proj_name
        else:
            add(1, 'IfcSite', site.GlobalId, site.Name, proj_name)
            path0 = f"{proj_name} > {site.Name}"
            buildings = []
            for rel in (site.IsDecomposedBy or []):
                buildings += [o for o in rel.RelatedObjects if o.is_a('IfcBuilding')]
            if not buildings:
                buildings = ifc_file.by_type('IfcBuilding')

        for building in buildings:
            path1 = f"{path0} > {building.Name}"
            add(2, 'IfcBuilding', building.GlobalId, building.Name, path0)

            storeys = []
            for rel in (building.IsDecomposedBy or []):
                storeys += [o for o in rel.RelatedObjects if o.is_a('IfcBuildingStorey')]
            storeys.sort(key=lambda s: (s.Elevation if s.Elevation is not None else 0))

            for st in storeys:
                path2 = f"{path1} > {st.Name}"
                add(3, 'IfcBuildingStorey', st.GlobalId, st.Name, path1, extra=f"Elevation={st.Elevation}")

                contained = []
                for rel in (st.ContainsElements or []):
                    contained += list(rel.RelatedElements)
                spaces = [c for c in contained if c.is_a('IfcSpace')]
                others = [c for c in contained if not c.is_a('IfcSpace')]

                for sp in spaces:
                    add(4, 'IfcSpace', sp.GlobalId, sp.Name, path2, extra=(sp.LongName or ''))
                if others:
                    add(4, '(요소 합계)', '', f"{len(others)}개 요소 직접 포함", path2,
                        extra=', '.join(sorted(set(o.is_a() for o in others))))

    return pd.DataFrame(rows)


# ===================================================================
# 2. 전체 속성 Long 포맷 (모든 IfcProduct x 모든 Attribute/Pset/Qto)
# ===================================================================

def _build_long_format(ifc_file):
    rows = []
    products = ifc_file.by_type('IfcProduct')

    for ent in products:
        cls = ent.is_a()
        storey = _get_storey(ent)
        storey_name = storey.Name if storey else None
        space_name = None if cls == 'IfcSpace' else _get_spaces_via_boundary(ifc_file, ent)
        material = _get_material(ent)
        guid = _safe_attr(ent, 'GlobalId')
        name = _safe_attr(ent, 'Name')

        for attr in CORE_ATTRS:
            v = _safe_attr(ent, attr)
            if v is None:
                continue
            rows.append({
                'IFC_Class': cls, 'GlobalId': guid, 'Name': name,
                '층(Storey)': storey_name, '공간(Space)': space_name, '재질(Material)': material,
                '구분': 'Attribute', 'Pset': '-', '속성명': attr, '값': v,
            })

        for k, v in _flatten_psets(ent).items():
            pset_name, prop_name = k.split('.', 1)
            kind = 'Qto' if pset_name.startswith('Qto_') else 'Pset'
            rows.append({
                'IFC_Class': cls, 'GlobalId': guid, 'Name': name,
                '층(Storey)': storey_name, '공간(Space)': space_name, '재질(Material)': material,
                '구분': kind, 'Pset': pset_name, '속성명': prop_name, '값': v,
            })

        if len(rows) > MAX_ROWS_LONG:
            break

    return pd.DataFrame(rows)


# ===================================================================
# 3. 클래스별 Wide(행렬) 포맷 - IFC 안에 실제 존재하는 클래스만 동적 생성
# ===================================================================

def _build_wide_sheets(ifc_file):
    products = ifc_file.by_type('IfcProduct')
    by_class = OrderedDict()
    for ent in products:
        cls = ent.is_a()
        if cls in EXCLUDE_FROM_WIDE:
            continue
        by_class.setdefault(cls, []).append(ent)

    # 계층 우선순위(대지/건물/층/공간 먼저, 그 다음 요소들은 인스턴스 많은 순)
    priority = ['IfcSite', 'IfcBuilding', 'IfcBuildingStorey', 'IfcSpace']
    ordered_classes = [c for c in priority if c in by_class]
    ordered_classes += sorted(
        [c for c in by_class if c not in priority],
        key=lambda c: -len(by_class[c])
    )

    sheets = OrderedDict()
    for cls in ordered_classes:
        ents = by_class[cls]
        rows = []
        for ent in ents:
            storey = _get_storey(ent)
            flat = _flatten_psets(ent)
            row = {
                'GlobalId': _safe_attr(ent, 'GlobalId'),
                'Name': _safe_attr(ent, 'Name'),
                'ObjectType': _safe_attr(ent, 'ObjectType'),
                'PredefinedType': _safe_attr(ent, 'PredefinedType'),
                'Tag': _safe_attr(ent, 'Tag'),
                '층(Storey)': storey.Name if storey else None,
                '공간(Space)': None if cls == 'IfcSpace' else _get_spaces_via_boundary(ifc_file, ent),
                '재질(Material)': _get_material(ent),
            }
            if cls in DIMENSION_TARGET_CLASSES:
                row.update(_dimension_columns(ent))
            if cls in AREA_TARGET_CLASSES:
                row.update(_area_columns(ent, flat))
            row.update(flat)
            rows.append(row)
        sheets[_label(cls)] = pd.DataFrame(rows)
    return sheets


def load_interest_set(spec_xlsx_path):
    """
    관심 속성 집합을 (Namespace, 속성명) 튜플 set으로 반환.

    두 가지 포맷을 자동 감지:
    1. 경량 spec.xlsx (spec_extractor.py 출력물)
       → '속성목록' 시트의 'Namespace', '속성명_IFC' 컬럼을 직접 읽음
    2. 원본 sBIM 명세 엑셀 (260312_sBIM정보별_체계_v8_3.xlsx 형식)
       → 각 시트에서 'IFC 스키마 (상세)' 컬럼을 파싱
    """
    wb = openpyxl.load_workbook(spec_xlsx_path, data_only=True)
    interest = set()

    # ── 포맷 1 감지: '속성목록' 시트 + 'Namespace'/'속성명_IFC' 컬럼 ──
    if '속성목록' in wb.sheetnames:
        ws = wb['속성목록']
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            i_ns   = header.index('Namespace') + 1
            i_prop = header.index('속성명_IFC') + 1
        except ValueError:
            i_ns, i_prop = None, None

        if i_ns and i_prop:
            for r in range(2, ws.max_row + 1):
                ns   = ws.cell(row=r, column=i_ns).value
                prop = ws.cell(row=r, column=i_prop).value
                if ns and prop:
                    interest.add((str(ns).strip(), str(prop).strip()))
            return interest

    # ── 포맷 2: 원본 명세 엑셀 — 시트별 'IFC 스키마' 컬럼 파싱 ──
    def _parse_token(token):
        token = token.strip()
        if not token or any(x in token for x in ['->', '(', ')']):
            return None
        if token.startswith('IfcRel') and '.' not in token:
            return None
        parts = token.split('.')
        if len(parts) < 2:
            return None
        ns, prop = (parts[-2].strip(), parts[-1].strip()) if len(parts) >= 3 \
                   else (parts[0].strip(), parts[1].strip())
        return (ns, prop) if ns and prop else None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, col_idx = None, None
        for r in range(1, min(ws.max_row + 1, 15)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and 'IFC' in v and '스키마' in v:
                    header_row, col_idx = r, c
                    break
            if header_row:
                break
        if header_row is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            raw = ws.cell(row=r, column=col_idx).value
            if not raw or not isinstance(raw, str):
                continue
            for eq_part in raw.split('='):
                for slash_part in eq_part.split('/'):
                    result = _parse_token(slash_part.strip())
                    if result:
                        interest.add(result)

    return interest


def _format_props_cell(flat_dict, interest_set=None):
    """{'Pset.속성':값, ...} -> 한 셀에 들어갈 멀티라인 텍스트.
    Pset 별로 묶어서 'PsetName' 줄바꿈 후 '  속성 = 값' 형태로 정리.
    interest_set이 주어지면, (Pset명, 속성명)이 거기 포함된 줄을 빨간색 굵게 표시한
    openpyxl CellRichText를 반환 (엑셀 셀 안에서 부분 색상 적용)."""
    if not flat_dict:
        return ''

    grouped = OrderedDict()
    for k, v in flat_dict.items():
        pset_name, prop_name = k.split('.', 1)
        grouped.setdefault(pset_name, []).append((prop_name, v))

    if not interest_set:
        lines = []
        for pset_name, props in grouped.items():
            lines.append(f"[{pset_name}]")
            for prop_name, v in props:
                v_str = '' if v is None else str(v)
                lines.append(f"  {prop_name} = {v_str}")
        return '\n'.join(lines)

    # interest_set이 있는 경우: 매칭되는 줄만 빨간 글씨로 강조한 rich text 구성
    black = InlineFont(color='000000')
    red = InlineFont(color='FF0000', b=True)
    blocks = []
    buf = ''  # 검정 텍스트를 모아뒀다가 하나의 TextBlock으로 묶기 위한 버퍼

    def flush_black():
        nonlocal buf
        if buf:
            blocks.append(TextBlock(black, buf))
            buf = ''

    for pset_name, props in grouped.items():
        buf += f"[{pset_name}]\n"
        for prop_name, v in props:
            v_str = '' if v is None else str(v)
            line = f"  {prop_name} = {v_str}\n"
            is_interest = (pset_name, prop_name) in interest_set
            if is_interest:
                flush_black()
                blocks.append(TextBlock(red, line))
            else:
                buf += line
    flush_black()

    if not blocks:
        return ''
    # 마지막 줄의 끝 개행 제거 (셀 끝 여백 방지)
    last = blocks[-1]
    if last.text.endswith('\n'):
        blocks[-1] = TextBlock(last.font, last.text[:-1])
    return CellRichText(*blocks)


# ===================================================================
# 3-1. 단일 셀(All-in-one cell) 통합 시트
#      계층/식별 정보는 컬럼(필터 가능), 모든 속성은 한 셀에 텍스트로 표기
# ===================================================================

def _build_consolidated_sheet(ifc_file, interest_set=None):
    products = ifc_file.by_type('IfcProduct')
    rows = []
    for ent in products:
        cls = ent.is_a()
        if cls in EXCLUDE_FROM_WIDE:
            continue
        storey = _get_storey(ent)
        flat = _flatten_psets(ent)
        base = {
            'IFC_Class': cls,
            '분류': _label(cls),
            '층(Storey)': storey.Name if storey else None,
            'Name': _safe_attr(ent, 'Name'),
            'GlobalId': _safe_attr(ent, 'GlobalId'),
            'ObjectType': _safe_attr(ent, 'ObjectType'),
            'PredefinedType': _safe_attr(ent, 'PredefinedType'),
            'Tag': _safe_attr(ent, 'Tag'),
            '재질(Material)': _get_material(ent),
        }
        if cls in DIMENSION_TARGET_CLASSES:
            base.update(_dimension_columns(ent))
        if cls in AREA_TARGET_CLASSES:
            base.update(_area_columns(ent, flat))
        base['속성개수'] = len(flat)
        base['전체속성(Pset.속성 = 값)'] = _format_props_cell(flat, interest_set)

        # 공간(Space)이 여러 개면 행을 분리 (1:N -> N개 행), 0개면 빈칸 1행, IfcSpace 자신은 해당없음
        space_list = [] if cls == 'IfcSpace' else _get_spaces_list_via_boundary(ifc_file, ent)
        if not space_list:
            row = dict(base)
            row['공간(Space)'] = None
            rows.append(row)
        else:
            for sp_name in space_list:
                row = dict(base)
                row['공간(Space)'] = sp_name
                rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        # 요청된 우선 순서: 층/공간/IFC_Class/분류/Name/치수XYZ/재질/속성개수/전체속성/치수산출방식
        priority_order = [
            '층(Storey)', '공간(Space)', 'IFC_Class', '분류', 'Name',
            '치수_X(m)', '치수_Y(m)', '치수_Z(m)', '면적(㎡)', '재질(Material)',
            '속성개수', '전체속성(Pset.속성 = 값)', '치수산출방식', '면적산출방식',
        ]
        col_order = [c for c in priority_order if c in df.columns]
        col_order += [c for c in df.columns if c not in col_order]  # 나머지(GlobalId 등)는 뒤에 부착
        df = df[col_order]
    return df


# ===================================================================
# 4. 엑셀 작성 (서식 포함)
# ===================================================================

HEADER_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
BODY_FONT = Font(name='맑은 고딕', size=10)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_df(ws, df):
    if df.empty:
        ws.append(['(데이터 없음)'])
        return
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    for _, r in df.iterrows():
        vals = []
        for v in r.tolist():
            if isinstance(v, (list, tuple)) and not isinstance(v, CellRichText):
                v = ', '.join(str(x) for x in v)
            vals.append(v)
        ws.append(vals)
    for rr in range(2, ws.max_row + 1):
        max_lines = 1
        for cc in range(1, len(df.columns) + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.font = BODY_FONT
            cell.border = BORDER
            n_lines = str(cell.value).count('\n') + 1 if cell.value else 1
            if n_lines > 1:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                max_lines = max(max_lines, n_lines)
            else:
                cell.alignment = Alignment(vertical='center')
        if max_lines > 1:
            ws.row_dimensions[rr].height = min(max_lines * 14, 400)
    sample = df.iloc[:200] if len(df) > 200 else df
    for c in range(1, len(df.columns) + 1):
        col_name = str(df.columns[c - 1])
        if '전체속성' in col_name or '속성정보' in col_name:
            ws.column_dimensions[get_column_letter(c)].width = 80
            continue
        maxlen = max(
            [len(str(df.columns[c - 1]))] +
            [len(str(v).split('\n')[0]) for v in sample.iloc[:, c - 1].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 2, 10), 45)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _unique_sheet_name(wb, name, used):
    base = name[:31]
    candidate = base
    i = 2
    while candidate in used:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


# ===================================================================
# 메인 함수
# ===================================================================

def extract_ifc_to_excel(ifc_path: str, output_path: str = None, include_long: bool = True,
                          consolidated: bool = True, per_class_wide: bool = True,
                          space_element_matrix: bool = True, spec_path: str = None,
                          naming_diagnosis: bool = False, naming_diagnosis_classes=None,
                          nvidia_api_key: str = None, nvidia_model: str = "meta/llama-3.1-70b-instruct",
                          status_cb=None) -> str:
    """
    임의의 IFC 파일을 읽어 계층 구조 기반 엑셀로 변환한다.

    Parameters
    ----------
    ifc_path : str
        입력 IFC 파일 경로
    output_path : str, optional
        출력 엑셀 경로. 미지정시 입력파일명 + '_추출.xlsx'
    include_long : bool, default True
        모든 속성을 한 줄씩 펼친 Long 포맷 시트 포함 여부 (대형 IFC의 경우 False 권장)
    consolidated : bool, default True
        계층 정보는 컬럼, 모든 속성은 '한 셀'에 모아 보여주는 통합 시트 포함 여부.
    per_class_wide : bool, default True
        클래스별로 속성을 컬럼으로 펼친 Wide 시트들 포함 여부 (벽/기둥/보 등 치수 컬럼 포함).
    space_element_matrix : bool, default True
        공간(Space)-부재 1:1 매칭 시트(04), 미매칭부재 시트(05), 벽 내외벽 판정 시트(06) 포함 여부.
        RelSpaceBoundary 1건 = 1행으로, 부재 1개가 여러 공간에 접하면 행이 분리되어 나타남.
        벽 내외벽 판정은 1차 Pset_WallCommon.IsExternal, 2차(1차 없을 때만) RelSpaceBoundary
        양면 Space 접촉 확인 순으로 수행한다.
    spec_path : str, optional
        검토 주관심사 명세 엑셀(예: sBIM 정보별 체계.xlsx) 경로. 지정하면 02_통합조회 시트의
        '전체속성' 셀에서 명세에 정의된 (Pset/Class, 속성) 에 해당하는 줄을 빨간 굵은 글씨로 강조.
    naming_diagnosis : bool, default False
        True면 NVIDIA NIM LLM으로 (IFC_Class, Name) 고유 조합이 IFC4 표준 의도와 부합하는지
        추론 진단해 '08_IFC4_명명규칙_진단' 시트를 추가한다. nvidia_api_key가 없으면 건너뛴다.
        주의: 이는 LLM 추론이며 확정된 오류 판정이 아니다(검토 참고용 의심목록).
    naming_diagnosis_classes : list[str], optional
        진단 대상 IFC_Class 제한 목록. None이면 전체 IfcElement 대상(기본값, 비용이 커질 수 있음).
    nvidia_api_key : str, optional
        NVIDIA NIM API 키. 라이브러리로 직접 호출할 때는 명시적으로 넘겨야 하며,
        CLI(main())에서만 환경변수 NVIDIA_API_KEY로 자동 fallback한다.
    nvidia_model : str, default "meta/llama-3.1-70b-instruct"
        사용할 NIM 모델명.
    status_cb : callable, optional
        str(message) -> None. 진행 상황(특히 LLM 배치 진행률) 콜백.

    Returns
    -------
    str : 생성된 엑셀 파일 경로
    """
    if output_path is None:
        output_path = ifc_path.rsplit('.', 1)[0] + '_추출.xlsx'

    ifc_file = ifcopenshell.open(ifc_path)

    interest_set = load_interest_set(spec_path) if spec_path else None

    cnt = Counter(e.is_a() for e in ifc_file)
    df_summary = pd.DataFrame(sorted(cnt.items(), key=lambda x: -x[1]), columns=['IFC_Class', '개수'])
    df_summary.loc[len(df_summary)] = ['IFC Schema', ifc_file.schema]
    if interest_set is not None:
        df_summary.loc[len(df_summary)] = ['관심속성(spec) 매칭개수', len(interest_set)]

    df_hierarchy = _build_hierarchy(ifc_file)

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    ws = wb.create_sheet(_unique_sheet_name(wb, '00_요약', used_names))
    _write_df(ws, df_summary)

    ws = wb.create_sheet(_unique_sheet_name(wb, '01_계층구조', used_names))
    _write_df(ws, df_hierarchy)

    if consolidated:
        df_consol = _build_consolidated_sheet(ifc_file, interest_set)
        ws = wb.create_sheet(_unique_sheet_name(wb, '02_통합조회(단일셀)', used_names))
        _write_df(ws, df_consol)

    if include_long:
        df_long = _build_long_format(ifc_file)
        ws = wb.create_sheet(_unique_sheet_name(wb, '03_전체속성(Long)', used_names))
        _write_df(ws, df_long)

    if space_element_matrix:
        wall_classification = _determine_wall_classification(ifc_file)

        df_matrix = _build_space_element_matrix(ifc_file, wall_classification=wall_classification)
        ws = wb.create_sheet(_unique_sheet_name(wb, '04_공간-부재_매칭(1대1)', used_names))
        _write_df(ws, df_matrix)

        df_unmatched = _build_unmatched_elements(ifc_file)
        ws = wb.create_sheet(_unique_sheet_name(wb, '05_미매칭부재(Space없음)', used_names))
        _write_df(ws, df_unmatched)

        df_wall_class = _build_wall_classification_sheet(ifc_file, wall_classification=wall_classification)
        ws = wb.create_sheet(_unique_sheet_name(wb, '06_벽_내외벽_판정', used_names))
        _write_df(ws, df_wall_class)

        df_dup = _build_relspaceboundary_duplication_report(ifc_file)
        ws = wb.create_sheet(_unique_sheet_name(wb, '07_RelSpaceBoundary_중복진단', used_names))
        _write_df(ws, df_dup)

    if per_class_wide:
        wide_sheets = _build_wide_sheets(ifc_file)
        for sheet_label, df in wide_sheets.items():
            ws = wb.create_sheet(_unique_sheet_name(wb, sheet_label, used_names))
            _write_df(ws, df)

    if naming_diagnosis:
        if not nvidia_api_key:
            print("[경고] naming_diagnosis=True이지만 nvidia_api_key가 없어 IFC4 명명규칙 진단을 건너뜁니다.", flush=True)
        else:
            llm_call = make_nim_llm_call(api_key=nvidia_api_key, model=nvidia_model)
            df_diag = run_ifc4_naming_diagnosis(
                ifc_file, llm_call, target_classes=naming_diagnosis_classes, status_cb=status_cb
            )
            ws = wb.create_sheet(_unique_sheet_name(wb, '08_IFC4_명명규칙_진단', used_names))
            _write_df(ws, df_diag)

    wb.save(output_path)
    return output_path


# ===================================================================
# CLI
# ===================================================================

def main():
    parser = argparse.ArgumentParser(description='IFC -> 계층형 행렬 엑셀 자동 추출')
    parser.add_argument('ifc_path', help='입력 IFC 파일 경로')
    parser.add_argument('output_path', nargs='?', default=None, help='출력 엑셀 경로 (생략 가능)')
    parser.add_argument('--no-long', action='store_true', help='전체속성(Long) 시트 생략 (대형 파일용)')
    parser.add_argument('--no-consolidated', action='store_true', help='통합조회(단일셀) 시트 생략')
    parser.add_argument('--no-wide', action='store_true', help='클래스별 행렬(Wide) 시트들 생략')
    parser.add_argument('--no-matrix', action='store_true', help='공간-부재 1:1 매칭/미매칭/내외벽판정 시트(04,05,06) 생략')
    parser.add_argument('--spec', default=None,
                         help='검토 주관심사 명세 엑셀 경로. 지정시 02 시트의 해당 속성을 빨간색으로 강조')
    args = parser.parse_args()

    # 참고: IFC4 명명규칙 LLM 진단(naming_diagnosis) 기능은 NVIDIA API 인증 이슈로 현재
    # CLI에서 비활성화된 상태입니다. 코드(run_ifc4_naming_diagnosis, make_nim_llm_call)는
    # 그대로 남아있으니, extract_ifc_to_excel(naming_diagnosis=True, nvidia_api_key=...)를
    # 직접 호출하면 재사용 가능합니다.

    out = extract_ifc_to_excel(
        args.ifc_path, args.output_path,
        include_long=not args.no_long,
        consolidated=not args.no_consolidated,
        per_class_wide=not args.no_wide,
        space_element_matrix=not args.no_matrix,
        spec_path=args.spec,
        naming_diagnosis=False,
    )
    print(f'완료: {out}')


if __name__ == '__main__':
    main()
